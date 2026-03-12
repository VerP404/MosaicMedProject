from __future__ import annotations

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.dn_reference.models import DnService, DnServicePrice, DnServicePricePeriod, DnServiceRequirement


def _norm(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


PERIOD_RE = re.compile(
    r"^с\s+(?P<start>\d{2}\.\d{2}\.\d{4})(?:\s+по\s+(?P<end>\d{2}\.\d{2}\.\d{4}))?$",
    re.IGNORECASE,
)


def _parse_date(value: str):
    return datetime.strptime(value, "%d.%m.%Y").date()


def _parse_price(value: object):
    raw = _norm(value).replace(" ", "").replace(",", ".")
    if not raw:
        return None
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


class Command(BaseCommand):
    help = "Импорт свода услуг ДН из Excel (№ п/п, Наименование услуги, Код услуги)."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Путь к Excel файлу «свод услуг.xlsx»")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (если не задано — первый)")
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Очистить текущие услуги и их связи перед импортом (осторожно!)",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["path"])
        sheet: Optional[str] = options["sheet"]
        clear = bool(options["clear"])

        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(
                "Для импорта нужен пакет openpyxl. Установите: pip install openpyxl"
            ) from e

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet] if sheet else wb.worksheets[0]

        # Ищем строку заголовков (№ п/п, Наименование услуги, Код услуги)
        header_row_idx = 1
        headers: list[str] = []
        headers_l: list[str] = []

        for r_idx in range(1, 11):
            row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx))[0]
            vals = [_norm(c.value) for c in row]
            vals_l = [v.lower() for v in vals]
            # Ищем ключевые слова в заголовках
            has_num = "№" in " ".join(vals) or "п/п" in " ".join(vals)
            has_name = "наименование" in " ".join(vals_l) and "услуг" in " ".join(vals_l)
            has_code = "код" in " ".join(vals_l) and "услуг" in " ".join(vals_l)
            if has_name and has_code:
                header_row_idx = r_idx
                headers = vals
                headers_l = vals_l
                break

        if not headers:
            row1 = list(ws.iter_rows(min_row=1, max_row=1))[0]
            headers = [_norm(c.value) for c in row1]
            headers_l = [h.lower() for h in headers]

        def col_idx(*variants: str) -> int:
            for v in variants:
                v_l = v.lower()
                for i, h in enumerate(headers_l):
                    if v_l in h:
                        return i
            raise CommandError(f"Не найдена колонка {variants} в заголовках: {headers}")

        try:
            i_num = col_idx("№", "п/п", "n", "номер", "no")
        except CommandError:
            i_num = 0
        try:
            i_name = col_idx("наименование", "услуг", "name", "название")
        except CommandError:
            i_name = 1
        try:
            i_code = col_idx("код", "услуг", "code")
        except CommandError:
            i_code = 2

        period_columns: list[tuple[int, tuple]] = []
        for idx, header in enumerate(headers):
            header_norm = _norm(header)
            match = PERIOD_RE.match(header_norm)
            if match:
                start = _parse_date(match.group("start"))
                end = _parse_date(match.group("end")) if match.group("end") else None
                period_columns.append((idx, (start, end, header_norm)))

        if clear:
            DnServiceRequirement.objects.all().delete()
            DnServicePrice.objects.all().delete()
            DnServicePricePeriod.objects.all().delete()
            DnService.objects.all().delete()
            self.stdout.write("Справочник услуг очищен.")

        created = 0
        updated = 0
        created_prices = 0
        order_val = 0

        for row in ws.iter_rows(min_row=header_row_idx + 1):
            values = [c.value for c in row]
            name = _norm(values[i_name] if i_name < len(values) else "")
            code = _norm(values[i_code] if i_code < len(values) else "")
            num_raw = values[i_num] if i_num < len(values) else None

            if not code:
                continue
            if not name:
                name = code

            order_val += 1
            try:
                order_val = int(num_raw) if num_raw is not None and str(num_raw).strip().isdigit() else order_val
            except (ValueError, TypeError):
                pass

            obj, created_flag = DnService.objects.update_or_create(
                code=code,
                defaults={"name": name, "order": order_val},
            )
            if created_flag:
                created += 1
            else:
                updated += 1

            for col_idx, (date_start, date_end, period_title) in period_columns:
                if col_idx >= len(values):
                    continue
                price = _parse_price(values[col_idx])
                if price is None:
                    continue
                period, _ = DnServicePricePeriod.objects.get_or_create(
                    date_start=date_start,
                    date_end=date_end,
                    defaults={"title": period_title, "is_active": date_end is None},
                )
                if period.title != period_title or period.is_active != (date_end is None):
                    period.title = period_title
                    period.is_active = date_end is None
                    period.save(update_fields=["title", "is_active"])

                _, price_created = DnServicePrice.objects.update_or_create(
                    service=obj,
                    period=period,
                    defaults={"price": price},
                )
                if price_created:
                    created_prices += 1

        wb.close()

        self.stdout.write(self.style.SUCCESS("Импорт свода услуг завершён."))
        self.stdout.write(
            f"  Создано услуг: {created}, обновлено услуг: {updated}, создано/обновлено цен: {created_prices}"
        )
