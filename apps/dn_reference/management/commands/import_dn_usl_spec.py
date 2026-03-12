from __future__ import annotations

from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.dn_reference.models import (
    DnDiagnosis,
    DnDiagnosisGroup,
    DnDiagnosisGroupMembership,
    DnService,
    DnServiceRequirement,
    DnSpecialty,
)


SKIP_PREFIXES = ("~$",)


def _norm(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _norm_code(value: object) -> str:
    return _norm(value).upper().replace(" ", "")


def _is_plus(value: object) -> bool:
    s = _norm(value)
    return s in {"+", "＋", "✓", "да", "Да", "YES", "yes", "1"}


def _expand_range_token(token: str) -> set[str]:
    token = token.replace("–", "-").replace("—", "-").replace(" ", "")
    if "-" not in token:
        return {token}

    left, right = token.split("-", 1)
    if not left or not right:
        return {token}

    def split_icd(code: str):
        code = code.upper()
        if not code:
            return None
        letter = code[0]
        rest = code[1:]
        if "." in rest:
            major, minor = rest.split(".", 1)
            if not major.isdigit() or not minor.isdigit():
                return None
            return letter, int(major), int(minor), True
        if not rest.isdigit():
            return None
        return letter, int(rest), None, False

    l = split_icd(left)
    r = split_icd(right)
    if not l or not r:
        return {token}
    if l[0] != r[0]:
        return {token}

    letter = l[0]
    if not l[3] and not r[3]:
        if l[1] > r[1]:
            return {token}
        return {f"{letter}{num:02d}" for num in range(l[1], r[1] + 1)}

    if l[1] == r[1] and l[3] and r[3]:
        if l[2] > r[2]:
            return {token}
        return {f"{letter}{l[1]:02d}.{num}" for num in range(l[2], r[2] + 1)}

    return {token}


def _rule_to_codes(rule: str) -> set[str]:
    codes: set[str] = set()
    for raw in rule.split(","):
        token = _norm_code(raw)
        if not token:
            continue
        codes.update(_expand_range_token(token))
    return codes


def _diagnosis_matches_rule(diagnosis_code: str, rule_codes: set[str]) -> bool:
    diagnosis_code = _norm_code(diagnosis_code)
    if not diagnosis_code:
        return False

    primary = diagnosis_code.split()[0]
    base = primary.split(".")[0]
    prefixes = {primary, base}

    if "." in primary:
        major, minor = primary.split(".", 1)
        prefixes.add(f"{major}.{minor}")

    return any(
        diagnosis_code.startswith(code) or primary.startswith(code) or base == code
        for code in rule_codes
    )


def _sync_group_memberships(group: DnDiagnosisGroup) -> int:
    rule_codes = _rule_to_codes(group.rule or group.code)
    if not rule_codes:
        return 0

    created = 0
    for diagnosis in DnDiagnosis.objects.all().only("id", "code"):
        if _diagnosis_matches_rule(diagnosis.code, rule_codes):
            _, was_created = DnDiagnosisGroupMembership.objects.get_or_create(
                group=group,
                diagnosis=diagnosis,
                defaults={"is_active": True},
            )
            if was_created:
                created += 1
    return created


def _detect_headers(ws):
    row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [_norm(v) for v in row]
    headers_l = [h.lower() for h in headers]

    def idx_by_contains(*variants: str, default: int | None = None):
        for i, h in enumerate(headers_l):
            if all(v.lower() in h for v in variants):
                return i
        return default

    name_idx = idx_by_contains("наименование", default=1)
    code_idx = idx_by_contains("код", default=2)
    if name_idx is None or code_idx is None:
        raise CommandError(f"Не найдены колонки услуги в заголовке: {headers}")

    group_cols: list[tuple[int, str]] = []
    for i, header in enumerate(headers):
        if i in (name_idx, code_idx):
            continue
        if header:
            group_cols.append((i, header))

    if not group_cols:
        raise CommandError(f"Не найдены группы диагнозов в заголовке: {headers}")

    return name_idx, code_idx, group_cols


class Command(BaseCommand):
    help = "Импорт файлов специальностей из папки usl_spec: один файл = одна специальность."

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            type=str,
            default=None,
            help="Папка с файлами специальностей (по умолчанию apps/dn_reference/data/usl_spec)",
        )
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Очистить старые требования услуг по специальностям перед загрузкой",
        )

    @transaction.atomic
    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("Установите openpyxl: pip install openpyxl") from e

        base = Path(options["path"]).resolve() if options["path"] else Path(__file__).resolve().parent.parent.parent / "data" / "usl_spec"
        if not base.exists():
            raise CommandError(f"Папка не найдена: {base}")

        if options["clear"]:
            DnServiceRequirement.objects.all().delete()
            DnDiagnosisGroupMembership.objects.all().delete()

        loaded = 0
        created_groups = 0
        created_requirements = 0
        created_memberships = 0

        for path in sorted(base.glob("*.xlsx")):
            if any(path.name.startswith(prefix) for prefix in SKIP_PREFIXES):
                self.stdout.write(f"Пропуск временного файла: {path.name}")
                continue

            specialty_name = path.stem.strip()
            self.stdout.write(f"Загрузка специальности: {specialty_name}")
            specialty, _ = DnSpecialty.objects.get_or_create(name=specialty_name)

            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]

            name_idx, code_idx, group_cols = _detect_headers(ws)
            groups_by_col = {}

            for order, (col_idx, group_header) in enumerate(group_cols, start=1):
                code = _norm_code(group_header)[:256]
                group, was_created = DnDiagnosisGroup.objects.get_or_create(
                    code=code,
                    defaults={
                        "title": group_header[:255],
                        "rule": group_header,
                        "order": order,
                    },
                )
                if was_created:
                    created_groups += 1
                else:
                    updated_fields = []
                    if not group.rule:
                        group.rule = group_header
                        updated_fields.append("rule")
                    if not group.title:
                        group.title = group_header[:255]
                        updated_fields.append("title")
                    if updated_fields:
                        group.save(update_fields=updated_fields)

                groups_by_col[col_idx] = group
                created_memberships += _sync_group_memberships(group)

            for row in ws.iter_rows(min_row=2, values_only=True):
                values = list(row)
                service_name = _norm(values[name_idx] if name_idx < len(values) else "")
                service_code = _norm(values[code_idx] if code_idx < len(values) else "")
                if not service_code:
                    continue
                if not service_name:
                    service_name = service_code

                service, _ = DnService.objects.get_or_create(
                    code=service_code,
                    defaults={"name": service_name},
                )
                if service.name != service_name:
                    service.name = service_name
                    service.save(update_fields=["name"])

                for col_idx, group in groups_by_col.items():
                    cell = values[col_idx] if col_idx < len(values) else None
                    if not _is_plus(cell):
                        continue
                    _, was_created = DnServiceRequirement.objects.get_or_create(
                        service=service,
                        group=group,
                        specialty=specialty,
                        defaults={"is_required": True},
                    )
                    if was_created:
                        created_requirements += 1

            wb.close()
            loaded += 1

        self.stdout.write(self.style.SUCCESS("Импорт файлов usl_spec завершен."))
        self.stdout.write(
            f"- обработано файлов: {loaded}\n"
            f"- создано групп: {created_groups}\n"
            f"- создано связей диагнозов с группами: {created_memberships}\n"
            f"- создано требований услуг: {created_requirements}"
        )

