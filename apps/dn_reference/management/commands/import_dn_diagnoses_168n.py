from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.dn_reference.models import (
    DnDiagnosis,
    DnDiagnosisCategory,
    DnDiagnosisSpecialty,
    DnSpecialty,
)


@dataclass(frozen=True)
class Row168n:
    ds: str
    speciality: str
    joint_speciality: str
    category: str


def _norm(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _mkb_norm(value: object) -> str:
    s = _norm(value).upper()
    if not s or s == "-":
        return ""
    # берем токен до пробела: "E11.9 сахарный диабет" -> "E11.9"
    return s.split()[0]


def _split_joint_specialties(value: str) -> list[str]:
    if not value:
        return []
    # В вашем файле встречается "терапия/ВОП" и похожие варианты
    parts = [p.strip() for p in value.split("/") if p.strip()]
    # убираем дубликаты с сохранением порядка
    seen = set()
    out: list[str] = []
    for p in parts:
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def _iter_rows_from_workbook(path: Path, sheet: Optional[str]) -> Iterable[Row168n]:
    try:
        import openpyxl
    except Exception as e:  # pragma: no cover
        raise CommandError(
            "Для импорта нужен пакет openpyxl. Установите его в окружение Django."
        ) from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    # Заголовки: ожидаем как минимум ds/speciality/joint_speciality/group (на вашем скрине именно так)
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_l = [h.lower() for h in header]

    def col_idx(*names: str) -> int:
        for n in names:
            n_l = n.lower()
            if n_l in header_l:
                return header_l.index(n_l)
        raise CommandError(f"Не найдена колонка {names} в заголовках: {header}")

    i_ds = col_idx("ds", "диагноз", "код", "мкб", "mkb")
    i_spec = col_idx("speciality", "specialty", "специальность")
    try:
        i_joint = col_idx("joint_speciality", "joint_specialty", "совместная специальность", "общая специальность")
    except CommandError:
        i_joint = None
    i_group = col_idx("group", "категория", "группа")

    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        yield Row168n(
            ds=_mkb_norm(values[i_ds] if i_ds < len(values) else ""),
            speciality=_norm(values[i_spec] if i_spec < len(values) else ""),
            joint_speciality=_norm(values[i_joint] if i_joint is not None and i_joint < len(values) else ""),
            category=_norm(values[i_group] if i_group < len(values) else ""),
        )


class Command(BaseCommand):
    help = "Импорт справочника диагнозов 168н (ds/speciality/joint_speciality/group) в модели ДН."

    def add_arguments(self, parser):
        parser.add_argument("path", type=str, help="Путь к Excel файлу 168н")
        parser.add_argument("--sheet", type=str, default=None, help="Имя листа (если не задано — первый)")
        parser.add_argument("--clear", action="store_true", help="Очистить текущие данные ДН-справочника перед импортом")

    @transaction.atomic
    def handle(self, *args, **options):
        path = Path(options["path"])
        sheet = options["sheet"]
        clear = bool(options["clear"])

        if not path.exists():
            raise CommandError(f"Файл не найден: {path}")

        if clear:
            # порядок важен из-за FK
            DnDiagnosisSpecialty.objects.all().delete()
            DnDiagnosis.objects.all().delete()
            DnSpecialty.objects.all().delete()
            DnDiagnosisCategory.objects.all().delete()

        created_diag = 0
        created_spec = 0
        created_cat = 0
        created_links = 0

        for r in _iter_rows_from_workbook(path, sheet):
            if not r.ds:
                continue

            cat_obj = None
            if r.category:
                cat_obj, cat_created = DnDiagnosisCategory.objects.get_or_create(name=r.category)
                if cat_created:
                    created_cat += 1

            diag_obj, diag_created = DnDiagnosis.objects.get_or_create(code=r.ds, defaults={"category": cat_obj})
            if diag_created:
                created_diag += 1
            else:
                # если категория появилась позже — аккуратно проставим
                if cat_obj and diag_obj.category_id is None:
                    diag_obj.category = cat_obj
                    diag_obj.save(update_fields=["category"])

            # основная специальность
            if r.speciality:
                spec_obj, spec_created = DnSpecialty.objects.get_or_create(name=r.speciality)
                if spec_created:
                    created_spec += 1
                _, link_created = DnDiagnosisSpecialty.objects.get_or_create(
                    diagnosis=diag_obj,
                    specialty=spec_obj,
                    source=DnDiagnosisSpecialty.SOURCE_PRIMARY,
                )
                if link_created:
                    created_links += 1

            # совместные специальности
            for joint in _split_joint_specialties(r.joint_speciality):
                spec_obj, spec_created = DnSpecialty.objects.get_or_create(name=joint)
                if spec_created:
                    created_spec += 1
                _, link_created = DnDiagnosisSpecialty.objects.get_or_create(
                    diagnosis=diag_obj,
                    specialty=spec_obj,
                    source=DnDiagnosisSpecialty.SOURCE_JOINT,
                )
                if link_created:
                    created_links += 1

        self.stdout.write(self.style.SUCCESS("Импорт 168н завершен."))
        self.stdout.write(
            f"- создано категорий: {created_cat}\n"
            f"- создано диагнозов: {created_diag}\n"
            f"- создано специальностей: {created_spec}\n"
            f"- создано связей диагноз<->специальность: {created_links}"
        )

