import csv
from datetime import date, timedelta
from datetime import datetime

from django.contrib import admin
from django.contrib.admin import SimpleListFilter
from django.http import HttpResponse
from django.urls import path, reverse
from django.contrib import messages
from django.utils.html import format_html
from django.shortcuts import render, redirect
from django.utils.translation import gettext_lazy as _
from django.http import HttpRequest
from import_export.admin import ImportExportModelAdmin
from unfold.admin import TabularInline, ModelAdmin
from unfold.decorators import action
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .forms import *
from .models import *
from .resource import DigitalSignatureResource, PostRG014Resource, SpecialtyRG014Resource
from ..data_loader.models.oms_data import *
from ..home.models import TelegramGroup
from ..organization.models import MiskauzDepartment, OMSDepartment


class DoctorRecordInline(TabularInline):
    model = DoctorRecord
    form = DoctorRecordForm
    extra = 0
    autocomplete_fields = ('specialty', 'profile')
    fields = ('doctor_code', 'start_date', 'end_date', 'department', 'specialty', 'profile',)
    verbose_name = "Запись врача"
    verbose_name_plural = "Записи врача"


class StaffRecordInline(TabularInline):
    model = StaffRecord
    extra = 0
    verbose_name = "Запись сотрудника"
    verbose_name_plural = "Записи сотрудников"


class RG014Inline(TabularInline):
    model = RG014
    form = RG014Form  # Используем форму с автозаполнением
    extra = 0
    fields = ('spec_issue_date', 'spec_name', 'cert_accred_sign', 'post_name', 'hire_date',
              'dismissal_date')
    verbose_name = "Запись RG014"
    verbose_name_plural = "Записи RG014"


class DigitalSignatureInline(TabularInline):
    model = DigitalSignature
    extra = 0
    fields = (
        'added_at', 'application_date', 'valid_from', 'valid_to', 'issued_date', 'certificate_serial', 'position',
        'scan', 'scan_uploaded_at',
        'revoked_date')
    readonly_fields = ('scan_uploaded_at', 'added_at')
    verbose_name = "ЭЦП"
    verbose_name_plural = "ЭЦП"


class MaternityLeaveInline(TabularInline):
    model = MaternityLeave
    extra = 0
    fields = ('start_date', 'end_date', 'note')
    verbose_name = "Декрет"
    verbose_name_plural = "Декреты"


class DigitalSignatureFilter(admin.SimpleListFilter):
    title = 'ЭЦП'
    parameter_name = 'digital_signature_status'

    def lookups(self, request, model_admin):
        queryset = model_admin.get_queryset(request)
        return (
            ('active', f'С ЭЦП ({self.get_count(queryset, True)})'),
            ('inactive', f'Без ЭЦП ({self.get_count(queryset, False)})'),
        )

    def get_count(self, queryset, is_active):
        today = date.today()
        if is_active:
            return queryset.filter(
                digital_signatures__valid_from__lte=today,
                digital_signatures__valid_to__gte=today
            ).distinct().count()
        return queryset.exclude(
            digital_signatures__valid_from__lte=today,
            digital_signatures__valid_to__gte=today
        ).distinct().count()

    def queryset(self, request, queryset):
        today = date.today()
        if self.value() == 'active':
            return queryset.filter(
                digital_signatures__valid_from__lte=today,
                digital_signatures__valid_to__gte=today
            ).distinct()
        elif self.value() == 'inactive':
            return queryset.exclude(
                digital_signatures__valid_from__lte=today,
                digital_signatures__valid_to__gte=today
            ).distinct()


class MaternityLeaveFilter(admin.SimpleListFilter):
    title = "В декрете"
    parameter_name = "is_on_maternity_leave"

    def lookups(self, request, model_admin):
        return (
            ('yes', "Да"),
            ('no', "Нет"),
        )

    def queryset(self, request, queryset):
        today = datetime.now().date()
        if self.value() == 'yes':
            return queryset.filter(
                maternity_leaves__start_date__lte=today
            ).filter(
                models.Q(maternity_leaves__end_date__isnull=True) |
                models.Q(maternity_leaves__end_date__gte=today)
            ).distinct()
        elif self.value() == 'no':
            # Исключаем записи с активными декретами
            return queryset.exclude(
                maternity_leaves__start_date__lte=today
            ).distinct()


class DigitalSignatureApplicationFilter(admin.SimpleListFilter):
    title = 'Состояние заявления'
    parameter_name = 'application_status'

    def lookups(self, request, model_admin):
        return (
            ('no_application', 'Нет заявления'),
            ('submitted', 'Заявление подано'),
        )

    def queryset(self, request, queryset):
        if self.value() == 'no_application':
            return queryset.filter(
                digital_signatures__scan__isnull=False,
                digital_signatures__application_date__isnull=True,
                digital_signatures__valid_from__isnull=True
            ).distinct()
        elif self.value() == 'submitted':
            return queryset.filter(
                digital_signatures__scan__isnull=False,
                digital_signatures__application_date__isnull=False,
                digital_signatures__valid_from__isnull=True
            ).distinct()


class DoctorCodeFilter(admin.SimpleListFilter):
    title = 'Фильтр по коду врача'
    parameter_name = 'doctor_code_filter'

    def lookups(self, request, model_admin):
        return (
            ('startswith', 'Начинается с'),
            ('contains', 'Содержит'),
        )

    def queryset(self, request, queryset):
        search_value = request.GET.get('q', '')  # Извлечение значения из строки поиска
        if not search_value:
            return queryset

        if self.value() == 'startswith':
            return queryset.filter(doctor_code__startswith=search_value)
        elif self.value() == 'contains':
            return queryset.filter(doctor_code__icontains=search_value)
        return queryset


# Форма для выбора отделения
class DepartmentForm(forms.Form):
    department = forms.CharField(label="Отделение", max_length=255)


@admin.register(DoctorRecord)
class DoctorRecordAdmin(ModelAdmin):
    change_list_template = "admin/doctor_change_list.html"
    list_display = ('person', 'doctor_code', 'start_date', 'end_date', 'department',)
    list_filter = ('start_date', 'end_date', 'department', DoctorCodeFilter)
    list_editable = ('start_date', 'end_date', 'department')

    # Настройка поиска по doctor_code (начинается с и содержит)
    search_fields = ('doctor_code', 'person__last_name', 'person__first_name', 'person__snils')
    actions = ['set_department', 'update_work_dates']

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('update-doctor-records/', self.admin_site.admin_view(self.update_doctor_records),
                 name='update_doctor_records'),
            path('insert-doctors-into-person/', self.admin_site.admin_view(self.insert_doctors_into_person),
                 name='insert_doctors_into_person'),
            path('match-departments/', self.admin_site.admin_view(self.match_departments), name='match_departments'),

        ]
        return custom_urls + urls

    def match_departments(self, request):
        updated_count = 0

        # Получаем все записи DoctorRecord
        doctor_records = DoctorRecord.objects.all()

        for record in doctor_records:
            # Сначала ищем совпадение в OMSDepartment
            oms_department = OMSDepartment.objects.filter(name=record.structural_unit).first()
            if oms_department:
                record.department = oms_department.department
                record.save()
                updated_count += 1
                continue  # Переходим к следующей записи

            # Если не найдено в OMSDepartment, ищем в MiskauzDepartment
            miskauz_department = MiskauzDepartment.objects.filter(name=record.structural_unit).first()
            if miskauz_department:
                record.department = miskauz_department.department
                record.save()
                updated_count += 1

        # Выводим сообщение об обновленных записях
        messages.success(request, f'Обновлено {updated_count} записей в DoctorRecord.')
        return redirect('admin:personnel_doctorrecord_changelist')

    def insert_doctors_into_person(self, request):
        inserted_count = 0
        existing_snils = []

        # Получаем все записи из DoctorData
        doctor_data_records = DoctorData.objects.all()

        for doctor_data in doctor_data_records:
            # Проверяем, существует ли Person с таким СНИЛС
            if not Person.objects.filter(snils=doctor_data.snils).exists():
                # Преобразуем дату рождения в формат YYYY-MM-DD
                try:
                    birth_date = datetime.strptime(doctor_data.birth_date, '%d-%m-%Y').date()
                except ValueError:
                    messages.error(request,
                                   f"Неверный формат даты для СНИЛС {doctor_data.snils}: {doctor_data.birth_date}")
                    continue

                # Если такого физлица нет, создаём новое
                Person.objects.create(
                    snils=doctor_data.snils,
                    last_name=doctor_data.last_name,
                    first_name=doctor_data.first_name,
                    patronymic=doctor_data.middle_name,
                    date_of_birth=birth_date,
                    gender=doctor_data.gender[0].upper()  # предполагается, что гендер представлен буквой (M/F)
                )
                inserted_count += 1
            else:
                existing_snils.append(doctor_data.snils)

        messages.success(request, f'Вставлено {inserted_count} новых записей в Person.')
        if existing_snils:
            messages.warning(request, f"Следующие СНИЛС уже существуют: {', '.join(existing_snils)}")

        return redirect('admin:personnel_doctorrecord_changelist')

    def update_doctor_records(self, request):
        updated_count = 0
        missing_snils = []

        # Получаем все записи из DoctorData
        doctor_data_records = DoctorData.objects.all()

        for doctor_data in doctor_data_records:
            try:
                # Ищем соответствующее лицо по СНИЛС
                person = Person.objects.get(snils=doctor_data.snils)

                # Проверка наличия записи в DoctorRecord для данного лица
                doctor_record, created = DoctorRecord.objects.get_or_create(
                    person=person,
                    doctor_code=doctor_data.doctor_code,
                )

                # Обновляем профиль и специальность, если они заданы
                if doctor_data.medical_profile_code:
                    # Извлекаем код профиля до пробела
                    profile_code = doctor_data.medical_profile_code.split(' ')[0]
                    try:
                        # Находим профиль по коду
                        profile = Profile.objects.get(code=profile_code)
                        doctor_record.profile = profile
                    except Profile.DoesNotExist:
                        messages.warning(request, f"Профиль с кодом {profile_code} не найден для врача {person}")

                if doctor_data.specialty_code:
                    # Извлекаем код специальности до пробела
                    specialty_code = doctor_data.specialty_code.split(' ')[0]
                    try:
                        # Находим специальность по коду
                        specialty = Specialty.objects.get(code=specialty_code)
                        doctor_record.specialty = specialty
                    except Specialty.DoesNotExist:
                        messages.warning(request,
                                         f"Специальность с кодом {specialty_code} не найдена для врача {person}")

                # Заполняем поле структурного подразделения (structural_unit)
                doctor_record.structural_unit = doctor_data.department  # Прямое заполнение текстом

                # Сохраняем запись
                doctor_record.save()
                updated_count += 1

            except Person.DoesNotExist:
                missing_snils.append(doctor_data.snils)

        # Вывод сообщений об успешном обновлении и предупреждениях
        messages.success(request, f'Обновлено {updated_count} записей врачей.')
        if missing_snils:
            messages.warning(request, f"СНИЛС отсутствуют в Person: {', '.join(missing_snils)}")

        return redirect('admin:personnel_doctorrecord_changelist')

    def update_work_dates(self, request, queryset):
        # Получаем все записи DoctorData
        doctor_data_records = DoctorData.objects.all()
        updated_count = 0

        for record in queryset:
            # Ищем записи в DoctorData по doctor_code и snils
            matching_records = doctor_data_records.filter(
                doctor_code=record.doctor_code,
                snils=record.person.snils
            )

            if not matching_records.exists():
                continue

            # Берем первую подходящую запись (если их несколько, нужно решить логику выбора)
            doctor_data = matching_records.first()

            # Обновляем start_date, если он пуст
            if not record.start_date and doctor_data.start_date != '-':
                try:
                    record.start_date = datetime.strptime(doctor_data.start_date, '%d-%m-%Y').date()
                except ValueError:
                    # Пропускаем, если формат даты некорректный
                    self.message_user(request, f"Некорректная дата начала работы для {doctor_data.snils}",
                                      level="error")
                    continue

            # Обновляем end_date, если указана дата в DoctorData
            if doctor_data.end_date != '-' and doctor_data.end_date:
                try:
                    record.end_date = datetime.strptime(doctor_data.end_date, '%d-%m-%Y').date()
                except ValueError:
                    # Пропускаем, если формат даты некорректный
                    self.message_user(request, f"Некорректная дата окончания работы для {doctor_data.snils}",
                                      level="error")
                    continue

            # Сохраняем обновления
            record.save()
            updated_count += 1

        self.message_user(request, f"Обновлено {updated_count} записей.")

    update_work_dates.short_description = "Обновить даты работы"

    # Функция для массового присвоения отделений
    def set_department(self, request, queryset):
        debug_info = []  # Список для отладочной информации

        if 'apply' in request.POST:
            debug_info.append("POST-запрос получен.")
            form = DepartmentActionForm(request.POST)

            if form.is_valid():
                debug_info.append("Форма валидна.")
                department = form.cleaned_data['department']
                try:
                    # Обновляем записи в базе данных
                    updated_count = queryset.update(department=department)
                    messages.success(request, f'{updated_count} записей было обновлено.')
                    debug_info.append(f"Успешно обновлено {updated_count} записей.")
                    return redirect('admin:personnel_doctorrecord_changelist')  # Редирект после успешного обновления
                except Exception as e:
                    messages.error(request, f'Ошибка при обновлении записей: {str(e)}')
                    debug_info.append(f"Ошибка при обновлении: {str(e)}")
            else:
                messages.error(request, 'Форма неверна, проверьте введённые данные.')
                debug_info.append("Форма не валидна.")
        else:
            debug_info.append("Форма не была отправлена.")
            form = DepartmentActionForm()

        # Возвращаем отладочную информацию на страницу
        return render(request, 'admin/set_department_action.html', {
            'form': form,
            'records': queryset,
            'debug_info': '\n'.join(debug_info)
        })


# Кастомный фильтр для отбора по Telegram группам
class TelegramGroupFilter(SimpleListFilter):
    title = 'Телеграм группы'
    parameter_name = 'telegram_group'

    def lookups(self, request, model_admin):
        groups = TelegramGroup.objects.all()
        return [(group.pk, group.name) for group in groups]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(telegram_groups__pk=self.value())
        return queryset


# Кастомный фильтр для статуса врача (Нет/Да/Уволен)
class DoctorStatusFilter(SimpleListFilter):
    title = 'Статус врача'
    parameter_name = 'doctor_status'

    def lookups(self, request, model_admin):
        return (
            ('none', 'Нет'),
            ('active', 'Да'),
            ('terminated', 'Уволен'),
        )

    def queryset(self, request, queryset):
        today = date.today()
        if self.value() == 'none':
            return queryset.filter(doctor_records__isnull=True)
        elif self.value() == 'active':
            return queryset.filter(doctor_records__isnull=False).filter(
                models.Q(doctor_records__end_date__isnull=True) | models.Q(doctor_records__end_date__gte=today)
            ).distinct()
        elif self.value() == 'terminated':
            # Пользователь имеет записи врача, но ни одна не активна
            return queryset.filter(doctor_records__isnull=False).exclude(
                models.Q(doctor_records__end_date__isnull=True) | models.Q(doctor_records__end_date__gte=today)
            ).distinct()
        return queryset


# Кастомный фильтр для статуса персонала
class StaffStatusFilter(SimpleListFilter):
    title = 'Статус персонала'
    parameter_name = 'staff_status'

    def lookups(self, request, model_admin):
        return (
            ('none', 'Нет'),
            ('active', 'Да'),
            ('terminated', 'Уволен'),
        )

    def queryset(self, request, queryset):
        today = date.today()
        if self.value() == 'none':
            return queryset.filter(staff_record__isnull=True)
        elif self.value() == 'active':
            return queryset.filter(staff_record__isnull=False).filter(
                models.Q(staff_record__end_date__isnull=True) | models.Q(staff_record__end_date__gte=today)
            )
        elif self.value() == 'terminated':
            return queryset.filter(staff_record__isnull=False).filter(
                staff_record__end_date__lt=today
            )
        return queryset


@admin.register(Person)
class PersonAdmin(ModelAdmin):
    list_display = (
        'fio_dob', 'is_doctor_display', 'is_staff_display', 'phone_number', 'telegram', 'digital_signature_status',
        'maternity_leave_status')
    search_fields = ('last_name', 'first_name', 'patronymic', 'snils', 'email', 'phone_number', 'telegram')
    list_filter = (
        DoctorStatusFilter, StaffStatusFilter, 'citizenship', TelegramGroupFilter, DigitalSignatureFilter, MaternityLeaveFilter,
        DigitalSignatureApplicationFilter)
    fieldsets = (
        ("Персональные данные", {
            "classes": ("four-columns",),
            'fields': ('last_name', 'first_name', 'patronymic', 'date_of_birth')
        }),
        (None, {
            "classes": ("three-columns",),
            'fields': ('snils', 'gender', 'citizenship')
        }),
        ('Контакты', {
            "classes": ("four-columns",),
            'fields': ('email', 'phone_number', 'telegram')
        }),
        ('Телеграм', {
            'fields': ('telegram_groups',)
        }),
    )
    inlines = [DoctorRecordInline, StaffRecordInline, RG014Inline, DigitalSignatureInline, MaternityLeaveInline]
    filter_horizontal = ('telegram_groups',)
    list_per_page = 15
    actions_list = ["export_expiring_digital_signatures"]

    def is_doctor_display(self, obj):
        qs = obj.doctor_records.all()
        today = date.today()
        if not qs.exists():
            return format_html('<span style="color: red;">Нет</span>')
        if qs.filter(models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)).exists():
            return format_html('<span style="color: green;">Да</span>')
        return format_html('<span style="color: orange;">Уволен</span>')

    is_doctor_display.short_description = "Врач"

    def is_staff_display(self, obj):
        if not hasattr(obj, 'staff_record'):
            return format_html('<span style="color: red;">Нет</span>')
        staff = obj.staff_record
        today = date.today()
        if staff.end_date is None or staff.end_date >= today:
            return format_html('<span style="color: green;">Да</span>')
        return format_html('<span style="color: orange;">Уволен</span>')

    is_staff_display.short_description = "Персонал"

    def get_readonly_fields(self, request, obj=None):
        if not request.user.is_superuser:
            return self.readonly_fields + ('telegram_groups',)
        return self.readonly_fields

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.prefetch_related('doctor_records', 'telegram_groups', 'maternity_leaves')
        return qs

    def digital_signature_status(self, obj):
        today = date.today()
        digital_signature = obj.digital_signatures.filter(valid_from__lte=today, valid_to__gte=today).first()
        if digital_signature:
            return format_html('<span style="color: green;">✔</span>')
        else:
            return format_html('<span style="color: red;">✘</span>')

    digital_signature_status.short_description = 'ЭЦП'

    def maternity_leave_status(self, obj):
        if obj.is_on_maternity_leave:
            return format_html('<span style="color: green;">✔</span>')
        return format_html('<span style="color: red;">✘</span>')

    maternity_leave_status.short_description = "В декрете"

    def fio_dob(self, obj):
        """
        Возвращает ФИО сотрудника и дату рождения (в формате дд.мм.гггг) в виде ссылки на change page.
        """
        dob = obj.date_of_birth.strftime("%d.%m.%Y") if obj.date_of_birth else ""
        # Формируем URL для перехода на страницу редактирования (пример; скорректируйте reverse, если нужно)
        url = reverse("admin:personnel_person_change", args=[obj.pk])
        return format_html("<a href='{}'>{} {}</a>", url, obj.__str__(), dob)

    fio_dob.short_description = "ФИО, ДР"

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

    @action(
        description=_("Экспорт сотрудников с истекающей ЭЦП"),
        url_path="export-expiring-digital-signatures",
        permissions=["export_expiring_digital_signatures"]
    )
    def export_expiring_digital_signatures(self, request: HttpRequest):
        """Экспорт сотрудников с истекающей ЭЦП в Excel"""
        if request.method == 'POST':
            form = ExpiringDigitalSignatureForm(request.POST)
            if form.is_valid():
                days = form.cleaned_data['days']
                return self._export_expiring_digital_signatures_to_excel(days)
        else:
            form = ExpiringDigitalSignatureForm(initial={'days': 30})

        context = {
            'form': form,
            'title': 'Экспорт сотрудников с истекающей ЭЦП',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/personnel/export_expiring_digital_signatures.html', context)

    def has_export_expiring_digital_signatures_permission(self, request: HttpRequest):
        """Проверка прав на экспорт"""
        return request.user.has_perm('personnel.export_expiring_digital_signatures') or request.user.is_superuser

    def _export_expiring_digital_signatures_to_excel(self, days: int):
        """Генерирует Excel файл с сотрудниками, у которых заканчивается ЭЦП"""
        today = date.today()
        expiration_date = today + timedelta(days=days)

        # Получаем всех сотрудников, которые работают сейчас
        # Проверяем через doctor_records и staff_records
        working_persons = Person.objects.filter(
            models.Q(
                # Активные врачи
                doctor_records__end_date__isnull=True
            ) | models.Q(
                doctor_records__end_date__gte=today
            ) | models.Q(
                # Активные сотрудники
                staff_records__end_date__isnull=True
            ) | models.Q(
                staff_records__end_date__gte=today
            )
        ).distinct().prefetch_related('doctor_records', 'staff_records', 'digital_signatures')

        # Собираем данные о сотрудниках с истекающей ЭЦП
        results = []
        for person in working_persons:
            # Проверяем, работает ли сотрудник сейчас
            is_doctor_active = False
            is_staff_active = False
            
            # Проверяем статус врача
            doctor_records = person.doctor_records.all()
            if doctor_records.exists():
                active_doctors = doctor_records.filter(
                    models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
                )
                if active_doctors.exists():
                    is_doctor_active = True
            
            # Проверяем статус персонала
            staff_records = person.staff_records.all()
            if staff_records.exists():
                active_staff = staff_records.filter(
                    models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
                )
                if active_staff.exists():
                    is_staff_active = True
            
            # Пропускаем, если сотрудник не работает
            if not is_doctor_active and not is_staff_active:
                continue

            # Получаем все действующие ЭЦП для сотрудника
            active_signatures = person.digital_signatures.filter(
                valid_from__lte=today,
                valid_to__gte=today,
                revoked_date__isnull=True
            ).order_by('-valid_to')

            if not active_signatures.exists():
                continue

            # Берем последнюю действующую ЭЦП (с самой поздней датой окончания)
            latest_signature = active_signatures.first()

            # Проверяем, истекает ли она в указанный период
            if latest_signature.valid_to <= expiration_date:
                # Определяем статус работы
                work_status = []
                if is_doctor_active:
                    work_status.append("Врач")
                if is_staff_active:
                    work_status.append("Персонал")
                work_status_str = ", ".join(work_status) if work_status else "Не определен"

                # Вычисляем количество дней до истечения
                days_until_expiration = (latest_signature.valid_to - today).days

                results.append({
                    'ФИО': str(person),
                    'СНИЛС': person.snils,
                    'ИНН': person.inn or '',
                    'Телефон': person.phone_number or '',
                    'Email': person.email or '',
                    'Телеграм': person.telegram or '',
                    'Статус работы': work_status_str,
                    'Серийный номер ЭЦП': latest_signature.certificate_serial or '',
                    'Должность': str(latest_signature.position) if latest_signature.position else '',
                    'Действует с': latest_signature.valid_from.strftime('%d.%m.%Y') if latest_signature.valid_from else '',
                    'Действует по': latest_signature.valid_to.strftime('%d.%m.%Y') if latest_signature.valid_to else '',
                    'Дней до истечения': days_until_expiration,
                })

        # Сортируем по количеству дней до истечения (сначала те, у кого меньше дней)
        results.sort(key=lambda x: x['Дней до истечения'])

        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Истекающие ЭЦП"

        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            '№', 'ФИО', 'СНИЛС', 'ИНН', 'Телефон', 'Email', 'Телеграм',
            'Статус работы', 'Серийный номер ЭЦП', 'Должность',
            'Действует с', 'Действует по', 'Дней до истечения'
        ]
        ws.append(headers)

        # Применяем стили к заголовкам
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Заполняем данные
        for idx, result in enumerate(results, start=1):
            row = [
                idx,
                result['ФИО'],
                result['СНИЛС'],
                result['ИНН'],
                result['Телефон'],
                result['Email'],
                result['Телеграм'],
                result['Статус работы'],
                result['Серийный номер ЭЦП'],
                result['Должность'],
                result['Действует с'],
                result['Действует по'],
                result['Дней до истечения'],
            ]
            ws.append(row)

            # Применяем стили к ячейкам
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Выделяем строки, где осталось меньше 7 дней
                if result['Дней до истечения'] < 7:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif result['Дней до истечения'] < 14:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

        # Настраиваем ширину столбцов
        column_widths = {
            'A': 6,   # №
            'B': 30,  # ФИО
            'C': 15,  # СНИЛС
            'D': 15,  # ИНН
            'E': 15,  # Телефон
            'F': 25,  # Email
            'G': 15,  # Телеграм
            'H': 15,  # Статус работы
            'I': 20,  # Серийный номер ЭЦП
            'J': 30,  # Должность
            'K': 15,  # Действует с
            'L': 15,  # Действует по
            'M': 18,  # Дней до истечения
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Фиксируем первую строку
        ws.freeze_panes = 'A2'

        # Создаем HTTP-ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"expiring_digital_signatures_{today.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


@admin.register(RG014)
class RG014Admin(ModelAdmin):
    list_display = (
        'person', 'organization', 'spec_issue_date', 'spec_name', 'cert_accred_sign',
        'post_name',
        'hire_date', 'dismissal_date')
    actions = ['export_as_csv']

    search_fields = ('person__last_name', 'person__first_name',)
    list_filter = ('spec_name', 'post_name', 'organization', 'cert_accred_sign')
    autocomplete_fields = ['organization', 'spec_name', 'post_name']
    ordering = ['person']

    def export_as_csv(self, request, queryset):
        meta = self.model._meta
        field_names = [field.name for field in meta.fields]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename={}.csv'.format(meta)
        writer = csv.writer(response)

        writer.writerow(field_names)
        for obj in queryset:
            row = writer.writerow([getattr(obj, field) for field in field_names])

        return response

    export_as_csv.short_description = "Export Selected"


class SpecialtyInline(TabularInline):
    model = Specialty
    extra = 1  # Количество пустых строк для добавления новых записей
    fields = ('code', 'description')  # Поля для редактирования


class ProfileInline(TabularInline):
    model = Profile
    extra = 1
    fields = ('code', 'description')


class SpecialtyRG014Inline(TabularInline):
    model = SpecialtyRG014
    extra = 1
    fields = ('code', 'description')


class PostRG014Inline(TabularInline):
    model = PostRG014
    extra = 1
    fields = ('code', 'description')


@admin.register(Specialty)
class SpecialtyAdmin(ModelAdmin):
    list_display = ('code', 'description')
    search_fields = ('code', 'description')


@admin.register(Profile)
class ProfileAdmin(ModelAdmin):
    list_display = ('code', 'description')
    search_fields = ('code', 'description')


@admin.register(SpecialtyRG014)
class SpecialtyRG014Admin(ImportExportModelAdmin, ModelAdmin):
    resource_class = SpecialtyRG014Resource
    list_display = ('code', 'description')
    search_fields = ('code', 'description')


@admin.register(PostRG014)
class PostRG014Admin(ImportExportModelAdmin, ModelAdmin):
    resource_class = PostRG014Resource
    list_display = ('code', 'description')
    search_fields = ('code', 'description')

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        extra_context['title'] = "Должности медицинских и фармацевтических работников (1.2.643.5.1.13.13.11.1002)"
        return super().changelist_view(request, extra_context=extra_context)


# Определяем кастомный фильтр
class StatusFilter(SimpleListFilter):
    title = 'Статус'
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('active', 'Действует'),
            ('inactive', 'Не действует'),
        )

    def queryset(self, request, queryset):
        today = date.today()
        if self.value() == 'active':
            return queryset.filter(valid_from__lte=today, valid_to__gte=today)
        if self.value() == 'inactive':
            return queryset.exclude(valid_from__lte=today, valid_to__gte=today)


class DigitalSignatureAdminForm(forms.ModelForm):
    class Meta:
        model = DigitalSignature
        fields = '__all__'

    def __init__(self, *args, **kwargs):
        super(DigitalSignatureAdminForm, self).__init__(*args, **kwargs)
        # Переопределяем ярлык для поля person на "Сотрудник"
        if 'person' in self.fields:
            self.fields['person'].label = "Сотрудник"
        # Отключаем редактирование для полей, отвечающих за дату загрузки и дату добавления, если они есть в форме
        if 'scan_uploaded_at' in self.fields:
            self.fields['scan_uploaded_at'].disabled = True
        if 'added_at' in self.fields:
            self.fields['added_at'].disabled = True


class DigitalSignatureStatusFilter(SimpleListFilter):
    title = 'Статус ЭЦП'  # Заголовок фильтра в админке
    parameter_name = 'ds_status'  # Параметр, по которому Django будет фильтровать

    def lookups(self, request, model_admin):
        """
        Возвращает кортежи (value, verbose_name),
        которые будут отображены в выпадающем списке фильтра.
        """
        return (
            ('active', 'Действующий'),
            ('revoked', 'Аннулирован'),
            ('expired', 'Истёк'),
            ('other', 'Прочие'),
        )

    def queryset(self, request, queryset):
        """
        Возвращает отфильтрованный queryset в зависимости от выбранного значения.
        """
        today = date.today()
        if self.value() == 'active':
            return queryset.filter(
                valid_from__lte=today,
                valid_to__gte=today,
                revoked_date__isnull=True
            )
        elif self.value() == 'revoked':
            return queryset.filter(revoked_date__isnull=False)
        elif self.value() == 'expired':
            return queryset.filter(
                valid_to__lt=today,
                revoked_date__isnull=True
            )
        elif self.value() == 'other':
            # Например, если valid_from или valid_to не заданы,
            # или любая другая нестандартная ситуация
            return queryset.filter(
                models.Q(valid_from__isnull=True) | models.Q(valid_to__isnull=True)
            )
        return queryset


@admin.register(DigitalSignature)
class DigitalSignatureAdmin(ImportExportModelAdmin, ModelAdmin):
    form = DigitalSignatureAdminForm
    resource_class = DigitalSignatureResource
    autocomplete_fields = ['person', 'position']
    readonly_fields = ('scan_uploaded_at', 'added_at')
    list_display = (
        'person',
        'valid_from',
        'valid_to',
        'certificate_serial',
        'position',
        'status',
        'has_scan'
    )
    list_filter = ('valid_from', 'valid_to', DigitalSignatureStatusFilter)
    search_fields = ('person__last_name', 'person__first_name', 'person__patronymic', 'person__snils')
    list_per_page = 20
    actions_list = ["export_expiring_digital_signatures"]
    fieldsets = (
        ("Основные данные", {
            "classes": ("section",),
            "fields": (
                ('person', 'position'),
                ('certificate_serial', 'application_date'),
            ),
        }),
        ("Даты действия", {
            "classes": ("section",),
            "fields": (
                ('valid_from', 'valid_to', 'issued_date', 'revoked_date'),
            ),
        }),
        ("Файл выписки", {
            "classes": ("section",),
            "fields": (
                ('scan', 'scan_uploaded_at', 'added_at'),
            ),
        }),
    )

    def has_import_permission(self, request):
        return False

    def get_import_formats(self):
        return []

    def status(self, obj):
        today = date.today()
        # 1) Если сертификат аннулирован или срок действия истёк:
        if obj.revoked_date or (obj.valid_to and obj.valid_to < today):
            return format_html('<span style="color: blue;">—</span>')
        # 2) Если действующий (сегодняшняя дата в пределах valid_from и valid_to):
        if obj.valid_from and obj.valid_to and obj.valid_from <= today <= obj.valid_to:
            return format_html('<span style="color: green;">✔</span>')
        # 3) Иначе считаем недействующим
        return format_html('<span style="color: red;">✘</span>')

    status.short_description = 'Статус'

    def has_scan(self, obj):
        if obj.scan:
            return format_html('<span style="color: green;">✔</span>')
        else:
            return format_html('<span style="color: red;">✘</span>')

    has_scan.short_description = 'Скан'

    def get_changeform_initial_data(self, request):
        """Предзаполнение полей из URL параметров"""
        initial = super().get_changeform_initial_data(request)
        
        # Предзаполнение сотрудника
        if 'person' in request.GET:
            try:
                person_id = int(request.GET.get('person'))
                initial['person'] = person_id
            except (ValueError, TypeError):
                pass
        
        # Предзаполнение должности
        if 'position' in request.GET:
            try:
                position_id = int(request.GET.get('position'))
                initial['position'] = position_id
            except (ValueError, TypeError):
                pass
        
        return initial

    @action(
        description=_("Экспорт сотрудников с истекающей ЭЦП"),
        url_path="export-expiring-digital-signatures",
        permissions=["export_expiring_digital_signatures"]
    )
    def export_expiring_digital_signatures(self, request: HttpRequest):
        """Экспорт сотрудников с истекающей ЭЦП в Excel"""
        if request.method == 'POST':
            form = ExpiringDigitalSignatureForm(request.POST)
            if form.is_valid():
                days = form.cleaned_data['days']
                return self._export_expiring_digital_signatures_to_excel(days)
        else:
            form = ExpiringDigitalSignatureForm(initial={'days': 30})

        context = {
            'form': form,
            'title': 'Экспорт сотрудников с истекающей ЭЦП',
            'opts': self.model._meta,
            'has_view_permission': self.has_view_permission(request),
        }
        return render(request, 'admin/personnel/export_expiring_digital_signatures.html', context)

    def has_export_expiring_digital_signatures_permission(self, request: HttpRequest):
        """Проверка прав на экспорт"""
        return request.user.has_perm('personnel.export_expiring_digital_signatures') or request.user.is_superuser

    def _export_expiring_digital_signatures_to_excel(self, days: int):
        """Генерирует Excel файл с сотрудниками, у которых заканчивается ЭЦП"""
        today = date.today()
        expiration_date = today + timedelta(days=days)

        # Получаем всех сотрудников, которые работают сейчас
        # Проверяем через doctor_records и staff_records
        working_persons = Person.objects.filter(
            models.Q(
                # Активные врачи
                doctor_records__end_date__isnull=True
            ) | models.Q(
                doctor_records__end_date__gte=today
            ) | models.Q(
                # Активные сотрудники
                staff_records__end_date__isnull=True
            ) | models.Q(
                staff_records__end_date__gte=today
            )
        ).distinct().prefetch_related('doctor_records', 'staff_records', 'digital_signatures')

        # Собираем данные о сотрудниках с истекающей ЭЦП
        results = []
        for person in working_persons:
            # Проверяем, работает ли сотрудник сейчас
            is_doctor_active = False
            is_staff_active = False
            
            # Проверяем статус врача
            doctor_records = person.doctor_records.all()
            if doctor_records.exists():
                active_doctors = doctor_records.filter(
                    models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
                )
                if active_doctors.exists():
                    is_doctor_active = True
            
            # Проверяем статус персонала
            staff_records = person.staff_records.all()
            if staff_records.exists():
                active_staff = staff_records.filter(
                    models.Q(end_date__isnull=True) | models.Q(end_date__gte=today)
                )
                if active_staff.exists():
                    is_staff_active = True
            
            # Пропускаем, если сотрудник не работает
            if not is_doctor_active and not is_staff_active:
                continue

            # Получаем все действующие ЭЦП для сотрудника
            active_signatures = person.digital_signatures.filter(
                valid_from__lte=today,
                valid_to__gte=today,
                revoked_date__isnull=True
            ).order_by('-valid_to')

            if not active_signatures.exists():
                continue

            # Берем последнюю действующую ЭЦП (с самой поздней датой окончания)
            latest_signature = active_signatures.first()

            # Проверяем, истекает ли она в указанный период
            if latest_signature.valid_to <= expiration_date:
                # Определяем статус работы
                work_status = []
                if is_doctor_active:
                    work_status.append("Врач")
                if is_staff_active:
                    work_status.append("Персонал")
                work_status_str = ", ".join(work_status) if work_status else "Не определен"

                # Вычисляем количество дней до истечения
                days_until_expiration = (latest_signature.valid_to - today).days

                results.append({
                    'ФИО': str(person),
                    'СНИЛС': person.snils,
                    'ИНН': person.inn or '',
                    'Телефон': person.phone_number or '',
                    'Email': person.email or '',
                    'Телеграм': person.telegram or '',
                    'Статус работы': work_status_str,
                    'Серийный номер ЭЦП': latest_signature.certificate_serial or '',
                    'Должность': str(latest_signature.position) if latest_signature.position else '',
                    'Действует с': latest_signature.valid_from.strftime('%d.%m.%Y') if latest_signature.valid_from else '',
                    'Действует по': latest_signature.valid_to.strftime('%d.%m.%Y') if latest_signature.valid_to else '',
                    'Дней до истечения': days_until_expiration,
                })

        # Сортируем по количеству дней до истечения (сначала те, у кого меньше дней)
        results.sort(key=lambda x: x['Дней до истечения'])

        # Создаем Excel файл
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Истекающие ЭЦП"

        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Заголовки
        headers = [
            '№', 'ФИО', 'СНИЛС', 'ИНН', 'Телефон', 'Email', 'Телеграм',
            'Статус работы', 'Серийный номер ЭЦП', 'Должность',
            'Действует с', 'Действует по', 'Дней до истечения'
        ]
        ws.append(headers)

        # Применяем стили к заголовкам
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Заполняем данные
        for idx, result in enumerate(results, start=1):
            row = [
                idx,
                result['ФИО'],
                result['СНИЛС'],
                result['ИНН'],
                result['Телефон'],
                result['Email'],
                result['Телеграм'],
                result['Статус работы'],
                result['Серийный номер ЭЦП'],
                result['Должность'],
                result['Действует с'],
                result['Действует по'],
                result['Дней до истечения'],
            ]
            ws.append(row)

            # Применяем стили к ячейкам
            for cell in ws[ws.max_row]:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                # Выделяем строки, где осталось меньше 7 дней
                if result['Дней до истечения'] < 7:
                    cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif result['Дней до истечения'] < 14:
                    cell.fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")

        # Настраиваем ширину столбцов
        column_widths = {
            'A': 6,   # №
            'B': 30,  # ФИО
            'C': 15,  # СНИЛС
            'D': 15,  # ИНН
            'E': 15,  # Телефон
            'F': 25,  # Email
            'G': 15,  # Телеграм
            'H': 15,  # Статус работы
            'I': 20,  # Серийный номер ЭЦП
            'J': 30,  # Должность
            'K': 15,  # Действует с
            'L': 15,  # Действует по
            'M': 18,  # Дней до истечения
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Фиксируем первую строку
        ws.freeze_panes = 'A2'

        # Создаем HTTP-ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"expiring_digital_signatures_{today.strftime('%Y%m%d')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response


@admin.register(DoctorReportingRecord)
class DoctorReportingRecordAdmin(ModelAdmin):
    list_display = ('person', 'fte', 'get_doctor_codes', 'building_name', 'department', 'start_date', 'end_date')
    search_fields = (
        'person__last_name', 'person__first_name',
        'department__name', 'doctor_records__doctor_code'
    )
    filter_horizontal = ['doctor_records']
    list_filter = ('department__building', 'department',)

    def building_name(self, obj):
        """
        Возвращает название корпуса, если у отдела есть связь с Building.
        """
        if obj.department and getattr(obj.department, 'building', None):
            return obj.department.building.name
        return "-"

    building_name.short_description = "Корпус"
    building_name.admin_order_field = 'department__building__name'

    def get_doctor_codes(self, obj):
        return ", ".join(dr.doctor_code for dr in obj.doctor_records.all())

    get_doctor_codes.short_description = "Коды врачей"

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "doctor_records":
            object_id = request.resolver_match.kwargs.get('object_id')
            if object_id:
                # Редактирование существующей записи
                instance = self.get_object(request, object_id)
                person = instance.person
                # Получаем все doctor_records для этого физлица,
                # которые уже назначены в других записях (исключая текущую)
                used_ids = DoctorReportingRecord.objects.filter(person=person) \
                    .exclude(pk=instance.pk) \
                    .values_list('doctor_records__pk', flat=True)
                # Выбираем doctor_records для данного Person, исключая те, что уже задействованы,
                # но включаем те, что уже выбраны в этой записи (чтобы они не пропали)
                qs = DoctorRecord.objects.filter(person=person).exclude(pk__in=used_ids) | instance.doctor_records.all()
                kwargs["queryset"] = qs.distinct()
            else:
                # При создании новой записи пытаемся получить идентификатор физлица из POST или GET
                person_id = request.POST.get('person') or request.GET.get('person')
                if person_id:
                    used_ids = DoctorReportingRecord.objects.filter(person_id=person_id) \
                        .values_list('doctor_records__pk', flat=True)
                    qs = DoctorRecord.objects.filter(person_id=person_id).exclude(pk__in=used_ids)
                    kwargs["queryset"] = qs
                else:
                    # Если физлицо не выбрано, можно вернуть пустой queryset
                    kwargs["queryset"] = DoctorRecord.objects.none()
        return super().formfield_for_manytomany(db_field, request, **kwargs)
