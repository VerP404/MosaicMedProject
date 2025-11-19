from datetime import datetime
from dash import html, dcc, Input, Output, State, exceptions, dash_table
import dash_bootstrap_components as dbc
from openpyxl import load_workbook
from io import BytesIO
import os
import base64

from apps.analytical_app.app import app
from apps.analytical_app.callback import TableUpdater
from apps.analytical_app.components.filters import (
    filter_years, filter_status, status_groups, update_buttons, filter_months
)
from apps.analytical_app.pages.economist.dispensary.query import (
    sql_query_oms_dispensary, sql_query_detailed_dispensary
)
from apps.analytical_app.query_executor import engine

type_page = "econ-dispensary-analysis"

# Путь к шаблону Excel
TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    'Диспансеризация.xlsx'
)


def fill_excel_template(oms_data, detailed_data, oms_columns, detailed_columns):
    """Заполняет шаблон Excel данными из запросов, удаляя умные таблицы и именованные диапазоны"""
    # Проверяем наличие шаблона
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Шаблон не найден: {TEMPLATE_PATH}")
    
    # Загружаем шаблон
    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    
    # Удаляем все именованные диапазоны, которые могут вызывать ошибки
    try:
        if hasattr(wb, 'defined_names') and wb.defined_names:
            # Создаем копию списка имен для безопасного удаления
            names_to_remove = list(wb.defined_names.definedName)
            for name in names_to_remove:
                try:
                    del wb.defined_names.definedName[name]
                except:
                    pass
    except Exception as e:
        print(f"Предупреждение: не удалось удалить именованные диапазоны: {e}")
    
    # Заполняем вкладку "ОМС"
    if 'ОМС' in wb.sheetnames:
        ws_oms = wb['ОМС']
        
        # Удаляем все умные таблицы с листа
        try:
            if hasattr(ws_oms, 'tables') and ws_oms.tables:
                tables_to_remove = list(ws_oms.tables.keys()) if isinstance(ws_oms.tables, dict) else []
                for table_name in tables_to_remove:
                    try:
                        del ws_oms.tables[table_name]
                    except:
                        pass
        except Exception as e:
            print(f"Предупреждение: не удалось удалить умные таблицы с листа ОМС: {e}")
        
        # Находим первую строку с данными (обычно строка 2, где заголовки в строке 1)
        header_row = 1
        start_data_row = 2
        
        # Определяем количество колонок из заголовков
        max_col = ws_oms.max_column
        
        # Очищаем данные ниже заголовков (начиная со 2 строки)
        if ws_oms.max_row >= start_data_row:
            ws_oms.delete_rows(start_data_row, ws_oms.max_row - start_data_row + 1)
        
        # Получаем названия колонок
        oms_column_names = [c["name"] if isinstance(c, dict) else c for c in oms_columns]
        
        # Записываем данные
        for row_idx, row_data in enumerate(oms_data, start=start_data_row):
            for col_idx, col_name in enumerate(oms_column_names, start=1):
                if col_idx <= max_col:
                    value = row_data.get(col_name, '')
                    ws_oms.cell(row=row_idx, column=col_idx, value=value)
    
    # Заполняем вкладку "Детализация"
    if 'Детализация' in wb.sheetnames:
        ws_detailed = wb['Детализация']
        
        # Удаляем все умные таблицы с листа
        try:
            if hasattr(ws_detailed, 'tables') and ws_detailed.tables:
                tables_to_remove = list(ws_detailed.tables.keys()) if isinstance(ws_detailed.tables, dict) else []
                for table_name in tables_to_remove:
                    try:
                        del ws_detailed.tables[table_name]
                    except:
                        pass
        except Exception as e:
            print(f"Предупреждение: не удалось удалить умные таблицы с листа Детализация: {e}")
        
        # Находим первую строку с данными (обычно строка 2, где заголовки в строке 1)
        header_row = 1
        start_data_row = 2
        
        # Определяем количество колонок из заголовков
        max_col = ws_detailed.max_column
        
        # Очищаем данные ниже заголовков (начиная со 2 строки)
        if ws_detailed.max_row >= start_data_row:
            ws_detailed.delete_rows(start_data_row, ws_detailed.max_row - start_data_row + 1)
        
        # Получаем названия колонок
        detailed_column_names = [c["name"] if isinstance(c, dict) else c for c in detailed_columns]
        
        # Записываем данные
        for row_idx, row_data in enumerate(detailed_data, start=start_data_row):
            for col_idx, col_name in enumerate(detailed_column_names, start=1):
                if col_idx <= max_col:
                    value = row_data.get(col_name, '')
                    ws_detailed.cell(row=row_idx, column=col_idx, value=value)
    
    # Сохраняем в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def economist_dispensary_analysis():
    return html.Div([
        # Фильтры
        dbc.Card(
            dbc.CardBody([
                dbc.CardHeader("Фильтры"),
                dbc.Row([
                    dbc.Col(update_buttons(type_page), width=2),
                    dbc.Col([
                        html.Div([
                            dbc.Button(
                                "Скачать Excel",
                                id=f'download-excel-button-{type_page}',
                                color="success",
                                className="mt-3"
                            ),
                            html.Div(
                                id=f'loading-status-{type_page}',
                                style={"display": "none", "margin-top": "10px", "text-align": "center"}
                            )
                        ])
                    ], width=2, style={"text-align": "center"}),
                    dbc.Col(filter_years(type_page), width=2),
                    dbc.Col([
                        html.Label("Отчетный месяц:"),
                        filter_months(type_page)
                    ], width=4),
                    dbc.Col([
                        html.Label("Статусы:"),
                        filter_status(type_page)
                    ], width=6),
                ], align="center"),
            ]),
            className="mb-3 shadow-sm"
        ),
        
        # Результаты в вкладках
        dbc.Tabs([
            dbc.Tab(
                label="ОМС",
                tab_id=f"tab-oms-{type_page}",
                children=[
                    dcc.Loading(
                        id=f'loading-oms-{type_page}',
                        type="default",
                        children=html.Div(id=f'result-oms-{type_page}')
                    )
                ]
            ),
            dbc.Tab(
                label="Детализация",
                tab_id=f"tab-detailed-{type_page}",
                children=[
                    dcc.Loading(
                        id=f'loading-detailed-{type_page}',
                        type="default",
                        children=html.Div(id=f'result-detailed-{type_page}')
                    )
                ]
            )
        ], active_tab=f"tab-oms-{type_page}"),
        
        # Компонент для скачивания файла
        dcc.Download(id=f'download-excel-{type_page}'),
        
        # Toast уведомления
        dbc.Toast(
            id=f'toast-excel-{type_page}',
            header="Уведомление",
            icon="info",
            duration=4000,
            is_open=False,
            dismissable=True,
            style={"position": "fixed", "top": 60, "right": 50, "width": 350},
        )
    ])


# Callback для переключения между групповым и индивидуальным выбором статусов
@app.callback(
    Output(f'status-group-container-{type_page}', 'style'),
    Output(f'status-individual-container-{type_page}', 'style'),
    Input(f'status-selection-mode-{type_page}', 'value')
)
def toggle_status_selection(mode):
    if mode == 'group':
        return {'display': 'block'}, {'display': 'none'}
    else:
        return {'display': 'none'}, {'display': 'block'}


# Callback для ОМС таблицы
@app.callback(
    Output(f'result-oms-{type_page}', 'children'),
    [Input(f'update-button-{type_page}', 'n_clicks')],
    [State(f'dropdown-year-{type_page}', 'value'),
     State(f'range-slider-month-{type_page}', 'value'),
     State(f'status-selection-mode-{type_page}', 'value'),
     State(f'status-group-radio-{type_page}', 'value'),
     State(f'status-individual-dropdown-{type_page}', 'value')]
)
def update_oms_table(n_clicks, year, month_range, status_mode, status_group, status_individual):
    if n_clicks is None:
        raise exceptions.PreventUpdate
    
    if not year or not month_range:
        return html.Div("Выберите год и месяц", className="text-muted")
    
    # Берем первый месяц из диапазона
    month = month_range[0] if isinstance(month_range, list) else month_range
    
    # Определяем статусы
    statuses = (
        status_groups.get(status_group, [])
        if status_mode == 'group' else (status_individual or [])
    )
    
    # Выполняем запрос
    query = sql_query_oms_dispensary(year, month, statuses)
    columns, data = TableUpdater.query_to_df(engine, query)
    
    if not data:
        return html.Div([
            dbc.Alert([
                html.H5("Данные не найдены", className="alert-heading"),
                html.P("По выбранным условиям не найдено ни одной записи.")
            ], color="info")
        ])
    
    return html.Div([
        dash_table.DataTable(
            id=f"table-oms-{type_page}",
            columns=[
                {
                    "name": c["name"] if isinstance(c, dict) else c,
                    "id": c["id"] if isinstance(c, dict) else c
                }
                for c in columns
            ],
            data=data,
            page_size=20,
            sort_action="native",
            filter_action="native",
            export_format="xlsx",
            style_table={"overflowX": "auto"},
        )
    ])


# Callback для Детализации таблицы
@app.callback(
    Output(f'result-detailed-{type_page}', 'children'),
    [Input(f'update-button-{type_page}', 'n_clicks')],
    [State(f'dropdown-year-{type_page}', 'value'),
     State(f'range-slider-month-{type_page}', 'value'),
     State(f'status-selection-mode-{type_page}', 'value'),
     State(f'status-group-radio-{type_page}', 'value'),
     State(f'status-individual-dropdown-{type_page}', 'value')]
)
def update_detailed_table(n_clicks, year, month_range, status_mode, status_group, status_individual):
    if n_clicks is None:
        raise exceptions.PreventUpdate
    
    if not year or not month_range:
        return html.Div("Выберите год и месяц", className="text-muted")
    
    # Берем первый месяц из диапазона
    month = month_range[0] if isinstance(month_range, list) else month_range
    
    # Определяем статусы
    statuses = (
        status_groups.get(status_group, [])
        if status_mode == 'group' else (status_individual or [])
    )
    
    # Выполняем запрос
    query = sql_query_detailed_dispensary(year, month, statuses)
    columns, data = TableUpdater.query_to_df(engine, query)
    
    if not data:
        return html.Div([
            dbc.Alert([
                html.H5("Данные не найдены", className="alert-heading"),
                html.P("По выбранным условиям не найдено ни одной записи.")
            ], color="info")
        ])
    
    return html.Div([
        dash_table.DataTable(
            id=f"table-detailed-{type_page}",
            columns=[
                {
                    "name": c["name"] if isinstance(c, dict) else c,
                    "id": c["id"] if isinstance(c, dict) else c
                }
                for c in columns
            ],
            data=data,
            page_size=20,
            sort_action="native",
            filter_action="native",
            export_format="xlsx",
            style_table={"overflowX": "auto"},
        )
    ])


# Callback для показа индикатора загрузки
@app.callback(
    [Output(f'loading-status-{type_page}', 'children'),
     Output(f'loading-status-{type_page}', 'style'),
     Output(f'download-excel-button-{type_page}', 'disabled')],
    [Input(f'download-excel-button-{type_page}', 'n_clicks')],
    prevent_initial_call=True
)
def show_loading_on_click(n_clicks):
    """Показывает индикатор загрузки при нажатии на кнопку"""
    if n_clicks:
        return (
            html.Div([
                dbc.Spinner(size="sm"),
                html.Span(" Формирование файла...", style={"margin-left": "10px"})
            ], style={"display": "flex", "align-items": "center", "justify-content": "center"}),
            {"display": "block", "margin-top": "10px"},
            True
        )
    return "", {"display": "none"}, False


# Callback для экспорта в Excel
@app.callback(
    [Output(f'download-excel-{type_page}', 'data'),
     Output(f'toast-excel-{type_page}', 'is_open', allow_duplicate=True),
     Output(f'toast-excel-{type_page}', 'children', allow_duplicate=True),
     Output(f'toast-excel-{type_page}', 'icon', allow_duplicate=True),
     Output(f'loading-status-{type_page}', 'children', allow_duplicate=True),
     Output(f'loading-status-{type_page}', 'style', allow_duplicate=True),
     Output(f'download-excel-button-{type_page}', 'disabled', allow_duplicate=True)],
    [Input(f'download-excel-button-{type_page}', 'n_clicks')],
    [State(f'dropdown-year-{type_page}', 'value'),
     State(f'range-slider-month-{type_page}', 'value'),
     State(f'status-selection-mode-{type_page}', 'value'),
     State(f'status-group-radio-{type_page}', 'value'),
     State(f'status-individual-dropdown-{type_page}', 'value')],
    prevent_initial_call=True
)
def export_to_excel(n_clicks, year, month_range, status_mode, status_group, status_individual):
    """Экспорт данных в Excel по шаблону"""
    if n_clicks is None:
        raise exceptions.PreventUpdate
    
    if not year or not month_range:
        return (
            None,
            True,
            "Выберите год и месяц для экспорта",
            "warning",
            "",
            {"display": "none"},
            False
        )
    
    try:
        # Берем первый месяц из диапазона
        month = month_range[0] if isinstance(month_range, list) else month_range
        
        # Определяем статусы
        statuses = (
            status_groups.get(status_group, [])
            if status_mode == 'group' else (status_individual or [])
        )
        
        # Выполняем запросы для получения данных
        oms_query = sql_query_oms_dispensary(year, month, statuses)
        detailed_query = sql_query_detailed_dispensary(year, month, statuses)
        
        oms_columns, oms_data = TableUpdater.query_to_df(engine, oms_query)
        detailed_columns, detailed_data = TableUpdater.query_to_df(engine, detailed_query)
        
        # Проверяем наличие данных
        if not oms_data and not detailed_data:
            return (
                None,
                True,
                "Нет данных для экспорта. Измените фильтры.",
                "warning",
                "",
                {"display": "none"},
                False
            )
        
        # Заполняем шаблон
        buffer = fill_excel_template(oms_data, detailed_data, oms_columns, detailed_columns)
        
        # Формируем имя файла
        filename = f"Диспансеризация_{year}_{month:02d}_{datetime.now():%Y%m%d_%H%M}.xlsx"
        
        # Получаем данные из буфера и кодируем в base64
        buffer.seek(0)
        excel_data = buffer.getvalue()
        b64 = base64.b64encode(excel_data).decode()
        
        # Возвращаем данные для скачивания и уведомление об успехе
        return (
            {
                "content": b64,
                "base64": True,
                "filename": filename,
                "type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            },
            True,
            f"✅ Файл '{filename}' готов к скачиванию",
            "success",
            "",
            {"display": "none"},
            False
        )
    
    except FileNotFoundError as e:
        return (
            None,
            True,
            f"Шаблон не найден: {TEMPLATE_PATH}",
            "danger",
            "",
            {"display": "none"},
            False
        )
    except Exception as e:
        # Логируем ошибку для отладки
        error_msg = str(e)
        import traceback
        print(f"Ошибка при экспорте в Excel: {error_msg}")
        print(f"Traceback: {traceback.format_exc()}")
        return (
            None,
            True,
            f"Ошибка при создании файла: {error_msg[:100]}",
            "danger",
            "",
            {"display": "none"},
            False
        )
