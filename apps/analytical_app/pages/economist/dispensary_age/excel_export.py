# -*- coding: utf-8 -*-
"""
Экспорт сводных ДР в Excel с объединёнными ячейками и оформлением.
"""
from io import BytesIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Стили
HEADER_FILL = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SUBTOTAL_FILL = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")
GRAND_FILL = PatternFill(start_color="CFE2FF", end_color="CFE2FF", fill_type="solid")
BOLD = Font(bold=True)
GROUP_FILL = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _fmt_cost(x):
    if x is None:
        return ""
    if isinstance(x, (int, float)):
        return f"{x:,.2f}".replace(",", " ").replace(".", ",")
    return str(x)


def _write_pivot_sheet(ws, rows, col_headers, group_col_idx=0, second_col_idx=1, skip_keys=("_row_type",)):
    """
    Пишет сводную на лист: объединение первой колонки по группам, стили для итогов.
    rows — список dict с ключами как в col_headers + _row_type.
    group_col_idx — индекс колонки группы (объединяем).
    """
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Разбиваем на блоки (как в HTML)
    blocks = []
    current_group = None
    current_block = []
    grand_total = None

    for r in rows:
        rt = r.get("_row_type")
        if rt == "detail":
            key = r.get(col_headers[group_col_idx], "")
            if key != current_group:
                if current_group is not None:
                    blocks.append((current_group, current_block))
                current_group = key
                current_block = []
            current_block.append(r)
        elif rt == "subtotal":
            current_block.append(r)
            blocks.append((current_group, current_block))
            current_group = None
            current_block = []
        elif rt == "grand_total":
            grand_total = r
            if current_group is not None:
                blocks.append((current_group, current_block))
            break

    row_num = 2
    for group_name, block_rows in blocks:
        start_row = row_num
        for j, row in enumerate(block_rows):
            is_subtotal = row.get("_row_type") == "subtotal"
            for c, key in enumerate(col_headers):
                if key in skip_keys:
                    continue
                val = row.get(key, "")
                cell = ws.cell(row=row_num, column=c + 1, value=val)
                cell.border = THIN_BORDER
                if is_subtotal:
                    cell.font = BOLD
                    cell.fill = SUBTOTAL_FILL
                if c == group_col_idx and j == 0:
                    cell.fill = GROUP_FILL
                    cell.font = Font(bold=True)
            row_num += 1
        # Объединение ячеек по группе (первая колонка)
        if start_row < row_num:
            ws.merge_cells(start_row=start_row, start_column=group_col_idx + 1, end_row=row_num - 1, end_column=group_col_idx + 1)
            ws.cell(row=start_row, column=group_col_idx + 1).alignment = Alignment(vertical="top", wrap_text=True)

    if grand_total:
        for c, key in enumerate(col_headers):
            if key in skip_keys:
                continue
            val = grand_total.get(key, "")
            cell = ws.cell(row=row_num, column=c + 1, value=val)
            cell.font = BOLD
            cell.fill = GRAND_FILL
            cell.border = THIN_BORDER
        if len(col_headers) >= 2:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=second_col_idx)
        row_num += 1

    # Ширина колонок
    for c in range(len(col_headers)):
        ws.column_dimensions[get_column_letter(c + 1)].width = max(14, min(50, len(str(col_headers[c])) + 2))


def _write_simple_sheet(ws, data: List[dict], sheet_title: str):
    """Простой лист: первая строка — заголовки из ключей, далее строки данных."""
    if not data:
        ws.cell(row=1, column=1, value="Нет данных")
        return
    keys = [k for k in data[0].keys() if not str(k).startswith("_")]
    for c, k in enumerate(keys, 1):
        cell = ws.cell(row=1, column=c, value=k)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for r_idx, row in enumerate(data, 2):
        for c, k in enumerate(keys, 1):
            v = row.get(k, "")
            cell = ws.cell(row=r_idx, column=c, value=v)
            cell.border = THIN_BORDER
    for c in range(len(keys)):
        ws.column_dimensions[get_column_letter(c + 1)].width = max(10, min(30, len(str(keys[c])) + 2))


def build_dr_excel(
    patient_costs: List[dict],
    summary: List[dict],
    combined_pivot: List[dict],
    pivot_dr1: List[dict],
    pivot_dr2: List[dict],
    type1: str,
    type2: str,
) -> BytesIO:
    """
    Собирает xlsx с листами: Стоимость по пациентам, Список type1/type2, сводные.
    type1, type2 — коды этапов (ДВ4/ДВ2, ДР1/ДР2, УД1/УД2).
    """
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    k_count1 = f"Количество талонов {type1}"
    k_count2 = f"Количество талонов {type2}"
    k_grp1 = f"Группа здоровья {type1}"
    k_grp2 = f"Группа здоровья {type2}"
    k_cost1 = f"Стоимость {type1}"
    k_cost2 = f"Стоимость {type2}"

    # 1. Стоимость по пациентам
    ws1 = wb.create_sheet("Стоимость по пациентам", 0)
    _write_simple_sheet(ws1, patient_costs, "Стоимость по пациентам")

    # 2. Список по первому этапу (количество талонов = 1)
    list1 = [r for r in patient_costs if (r.get(k_count1) or 0) == 1]
    ws_list1 = wb.create_sheet(f"Список {type1}", 1)
    _write_simple_sheet(ws_list1, list1, f"Список {type1}")

    # 3. Список по второму этапу (количество талонов = 1)
    list2 = [r for r in patient_costs if (r.get(k_count2) or 0) == 1]
    ws_list2 = wb.create_sheet(f"Список {type2}", 2)
    _write_simple_sheet(ws_list2, list2, f"Список {type2}")

    # 4. Сводная по двум этапам
    ws2 = wb.create_sheet(f"Сводная {type1} и {type2}", 3)
    headers_combined = [
        k_grp1, k_grp2, "Ж", "М", "Общий итог", k_cost1, k_cost2, "Общая стоимость",
    ]
    if combined_pivot:
        combined_for_excel = []
        for r in combined_pivot:
            r2 = dict(r)
            r2[k_cost1] = _fmt_cost(r.get(k_cost1))
            r2[k_cost2] = _fmt_cost(r.get(k_cost2))
            r2["Общая стоимость"] = _fmt_cost(r.get("Общая стоимость"))
            combined_for_excel.append(r2)
        _write_pivot_sheet(ws2, combined_for_excel, headers_combined, group_col_idx=0, second_col_idx=2)
    else:
        ws2.cell(row=1, column=1, value="Нет данных")

    # 5. Сводная по первому этапу
    ws3 = wb.create_sheet(f"Сводная {type1}", 4)
    headers_dr = ["Группа здоровья", "Стоимость", "Ж", "М", "Общий итог"]
    if pivot_dr1:
        _write_pivot_sheet(ws3, pivot_dr1, headers_dr, group_col_idx=0, second_col_idx=2)
    else:
        ws3.cell(row=1, column=1, value="Нет данных")

    # 6. Сводная по второму этапу
    ws4 = wb.create_sheet(f"Сводная {type2}", 5)
    if pivot_dr2:
        _write_pivot_sheet(ws4, pivot_dr2, headers_dr, group_col_idx=0, second_col_idx=2)
    else:
        ws4.cell(row=1, column=1, value="Нет данных")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
