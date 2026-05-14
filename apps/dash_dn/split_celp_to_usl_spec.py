"""
Разбивка CELP (Приложение по соответствию услуги и диагноза) в набор Excel
файлов формата usl_spec для load_usl_spec_dir / build_catalog_from_excel.

Пример:
  set PYTHONPATH=%CD%
  py -m apps.dash_dn.split_celp_to_usl_spec
  py -m apps.dash_dn.build_catalog_from_excel --usl-spec apps/dash_dn/usl_spec ...

Колонки «Тариф», «Пол», «№ п/п» в выходные матрицы не попадают (не группы МКБ).

Лист PROFIL=97,57,42 (ВОП / терапевт участковый / терапевт): одна матрица на троих;
строка с кодом B04.047.001/B04.047.003/B04.026.001 дублируется в три файла с одним кодом
и наименованием по НМУ для каждой специальности. Листы PROFIL=29 и PROFIL=60 сохраняются
как «Кардиолог» и «Онколог».
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Optional

_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
_DEFAULT_INPUT = Path(__file__).resolve().parent / (
    "CELP_3 Приложение по соответствию услуги и диагноза +тариф с 01.04.2026(1).xlsx"
)
_DEFAULT_OUT = Path(__file__).resolve().parent / "usl_spec"

# Лист с тремя профилями (97 / 57 / 42): одна матрица, в коде услуг — три варианта через «/».
PROFIL_TRIPLE_SHEET = "PROFIL=97,57,42"
TRIPLE_DISPENSARY_CODE_CANON = "B04.047.001/B04.047.003/B04.026.001"
# Порядок как в приказе: ВОП → B04.026.001; участковый терапевт → B04.047.003; терапевт → B04.047.001
TRIPLE_SPECIALTIES_ORDER = [
    "Врач общей практики",
    "Терапевт участковый",
    "Терапевт",
]
TRIPLE_DISPENSARY_BY_SPECIALTY: dict[str, tuple[str, str]] = {
    "Врач общей практики": (
        "B04.026.001",
        "Диспансерный прием (осмотр, консультация) врача по общей практике (семейного врача)",
    ),
    "Терапевт участковый": (
        "B04.047.003",
        "Диспансерный прием (осмотр, консультация) врача-терапевта участкового",
    ),
    "Терапевт": (
        "B04.047.001",
        "Диспансерный прием (осмотр, консультация) врача-терапевта",
    ),
}

# Человекочитаемые имена файлов вместо имён листов PROFIL=…
PROFIL_SHEET_OUTPUT_STEM: dict[str, str] = {
    "PROFIL=29": "Кардиолог",
    "PROFIL=60": "Онколог",
}

# Старые имена файлов после переименования — удаляем при успешной генерации.
OBSOLETE_USL_SPEC_NAMES = frozenset(
    {"PROFIL=29.xlsx", "PROFIL=60.xlsx", "PROFIL=97,57,42.xlsx"}
)

_WIN_BAD = re.compile(r'[<>:"/\\|?*]')


def _norm(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _norm_header_key(h: str) -> str:
    return _norm(h).lower().replace("ё", "е")


def _is_skip_service_meta_header(h: str) -> bool:
    k = _norm_header_key(h)
    if not k:
        return True
    if k in ("тариф", "пол", "пол пациента"):
        return True
    if k.startswith("№") or "п/п" in k:
        return True
    if k in ("n", "no", "номер", "ном."):
        return True
    return False


def _sanitize_filename(stem: str) -> str:
    s = _WIN_BAD.sub("_", stem).strip()
    s = s.rstrip(". ")
    return s or "unnamed"


def _norm_slash_service_code(code: str) -> str:
    """Сравнение составных кодов вида B04.../B04.../B04..."""
    s = _norm(code).upper().replace(" ", "").replace("\\", "/")
    parts = [p for p in s.split("/") if p]
    return "/".join(parts)


def _is_triple_dispensary_row(code: str) -> bool:
    return _norm_slash_service_code(code) == TRIPLE_DISPENSARY_CODE_CANON


def _apply_triple_dispensary_row(row: list[Any], specialty: str) -> list[Any]:
    """Дублирует строку матрицы с подстановкой одного кода/наименования для профиля."""
    out = list(row)
    code = _norm(out[1])
    if not _is_triple_dispensary_row(code):
        return out
    new_code, new_name = TRIPLE_DISPENSARY_BY_SPECIALTY[specialty]
    out[0] = new_name
    out[1] = new_code
    return out


def _detect_name_code_indices(headers: list[str]) -> tuple[int, int]:
    """0-based indices; aligned with _detect_usl_headers in build_catalog_from_excel."""
    headers_l = [_norm_header_key(h) for h in headers]

    def first_idx(substr: str) -> Optional[int]:
        sub = substr.lower()
        for i, h in enumerate(headers_l):
            if sub in h:
                return i
        return None

    name_i = first_idx("наименование")
    if name_i is None:
        name_i = 1 if len(headers) > 1 else 0
    code_i = first_idx("код")
    if code_i is None:
        code_i = 2 if len(headers) > 2 else (0 if name_i != 0 else 1)
    if code_i == name_i:
        for i, h in enumerate(headers_l):
            if i != name_i and "код" in h:
                code_i = i
                break
    return name_i, code_i


def _group_column_indices(headers: list[str], name_i: int, code_i: int) -> list[int]:
    out: list[int] = []
    for i, h in enumerate(headers):
        if i in (name_i, code_i):
            continue
        if _is_skip_service_meta_header(h):
            continue
        if not _norm(h):
            continue
        out.append(i)
    return out


def _write_matrix(
    out_path: Path,
    header_row: list[str],
    data_rows: list[list[Any]],
) -> None:
    import openpyxl

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(header_row)
    for row in data_rows:
        ws.append(row)
    wb.save(out_path)


def export_simple_matrix_sheet(
    ws,
    *,
    data_start_row: int,
    name_col_1based: int,
    code_col_1based: int,
    group_cols_1based: list[int],
    max_row: int,
) -> tuple[list[str], list[list[Any]]]:
    header_row = [
        "Наименование услуги",
        "Код услуги",
    ]
    for gc in group_cols_1based:
        h = _norm(ws.cell(1, gc).value)
        header_row.append(h or f"COL_{gc}")

    data_rows: list[list[Any]] = []
    for r in range(data_start_row, max_row + 1):
        name = _norm(ws.cell(r, name_col_1based).value)
        code = _norm(ws.cell(r, code_col_1based).value)
        if not code:
            continue
        if not name:
            name = code
        row_out: list[Any] = [name, code]
        for gc in group_cols_1based:
            row_out.append(ws.cell(r, gc).value)
        data_rows.append(row_out)
    return header_row, data_rows


def split_profil_like_sheet(ws, *, max_row: Optional[int] = None) -> tuple[list[str], list[list[Any]]]:
    max_col = ws.max_column or 0
    mr = max_row if max_row is not None else (ws.max_row or 0)
    headers = [_norm(ws.cell(1, c).value) for c in range(1, max_col + 1)]
    name_i, code_i = _detect_name_code_indices(headers)
    group_is = _group_column_indices(headers, name_i, code_i)
    g_1b = [i + 1 for i in group_is]
    return export_simple_matrix_sheet(
        ws,
        data_start_row=2,
        name_col_1based=name_i + 1,
        code_col_1based=code_i + 1,
        group_cols_1based=g_1b,
        max_row=mr,
    )


def _data_end_row_profil(ws, *, data_start_row: int = 2) -> int:
    """Последняя строка с кодом услуги (Axx/Bxx… или составной через /); хвост легенды отбрасывается."""
    max_row = ws.max_row or 0
    headers = [_norm(ws.cell(1, c).value) for c in range(1, (ws.max_column or 0) + 1)]
    name_i, code_i = _detect_name_code_indices(headers)
    code_1b = code_i + 1
    code_re = re.compile(r"^[ABab]\d{2}\.", re.I)
    end = data_start_row - 1
    for r in range(data_start_row, max_row + 1):
        code = _norm(ws.cell(r, code_1b).value)
        if not code or code == "+":
            continue
        if code_re.match(code) or "/" in code:
            end = r
    return max(end, data_start_row)


def split_prochee_sheet(ws) -> dict[str, tuple[list[str], list[list[Any]]]]:
    """ПРОЧИЕ: колонки A–E служебные; с 6-й — блоки, строка 2 специальность (forward-fill), строка 3 правило."""
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0
    cur_spec: Optional[str] = None
    spec_cols: dict[str, list[tuple[int, str]]] = {}

    for c in range(6, max_col + 1):
        v2 = ws.cell(2, c).value
        if v2 and str(v2).strip():
            cur_spec = str(v2).strip()
        v3 = ws.cell(3, c).value
        if cur_spec and v3 and str(v3).strip():
            rule = str(v3).strip()
            spec_cols.setdefault(cur_spec, []).append((c, rule))

    out: dict[str, tuple[list[str], list[list[Any]]]] = {}
    for spec, cols_rules in spec_cols.items():
        cols_rules.sort(key=lambda x: x[0])
        cols = [c for c, _ in cols_rules]
        rules = [r for _, r in cols_rules]
        header_row = ["Наименование услуги", "Код услуги", *rules]
        data_rows: list[list[Any]] = []
        for r in range(4, max_row + 1):
            name = _norm(ws.cell(r, 2).value)
            code = _norm(ws.cell(r, 3).value)
            if not code:
                continue
            if not name:
                name = code
            row_out: list[Any] = [name, code]
            for c in cols:
                row_out.append(ws.cell(r, c).value)
            data_rows.append(row_out)
        out[spec] = (header_row, data_rows)
    return out


def run(*, input_path: Path, out_dir: Path, dry_run: bool) -> list[Path]:
    import openpyxl

    if not input_path.is_file():
        raise FileNotFoundError(input_path)

    written: list[Path] = []
    used_names: dict[str, int] = {}

    def alloc_path(stem: str) -> Path:
        base = _sanitize_filename(stem)
        n = used_names.get(base, 0) + 1
        used_names[base] = n
        if n > 1:
            base = f"{base}_{n}"
        return out_dir / f"{base}.xlsx"

    wb = openpyxl.load_workbook(input_path, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if sheet_name.strip().upper() == "ПРОЧИЕ" or sheet_name == "ПРОЧИЕ":
                bundles = split_prochee_sheet(ws)
                for spec, (hdr, rows) in bundles.items():
                    path = alloc_path(spec)
                    if dry_run:
                        print(f"[dry-run] {path.name} <- ПРОЧИЕ/{spec} groups={len(hdr) - 2} data={len(rows)}")
                    else:
                        _write_matrix(path, hdr, rows)
                        written.append(path)
                        print(f"OK {path.name} (ПРОЧИЕ · {spec} · {len(rows)} строк)")
                continue

            if sheet_name == PROFIL_TRIPLE_SHEET:
                end_r = _data_end_row_profil(ws)
                hdr, rows = split_profil_like_sheet(ws, max_row=end_r)
                for spec in TRIPLE_SPECIALTIES_ORDER:
                    rows_spec = [_apply_triple_dispensary_row(list(r), spec) for r in rows]
                    path = alloc_path(spec)
                    if dry_run:
                        n_trip = sum(1 for r in rows if _is_triple_dispensary_row(_norm(r[1])))
                        print(
                            f"[dry-run] {path.name} <- {sheet_name}/{spec} "
                            f"groups={len(hdr) - 2} data={len(rows_spec)} "
                            f"(строк с тройным кодом дисп. приёма в исходнике: {n_trip})"
                        )
                    else:
                        _write_matrix(path, hdr, rows_spec)
                        written.append(path)
                        print(f"OK {path.name} ({sheet_name} → {spec} · {len(rows_spec)} строк)")
                continue

            end_r = _data_end_row_profil(ws)
            hdr, rows = split_profil_like_sheet(ws, max_row=end_r)
            stem = PROFIL_SHEET_OUTPUT_STEM.get(sheet_name, sheet_name)
            path = alloc_path(stem)
            if dry_run:
                print(f"[dry-run] {path.name} <- {sheet_name} groups={len(hdr) - 2} data={len(rows)}")
            else:
                _write_matrix(path, hdr, rows)
                written.append(path)
                print(f"OK {path.name} ({sheet_name} · {len(rows)} строк)")
    finally:
        wb.close()

    if not dry_run and written:
        for fn in OBSOLETE_USL_SPEC_NAMES:
            p = out_dir / fn
            if p.is_file():
                p.unlink()
                print(f"Удалён устаревший файл: {fn}")

    return written


def main() -> None:
    p = argparse.ArgumentParser(description="CELP xlsx → usl_spec/*.xlsx")
    p.add_argument("--input", type=Path, default=_DEFAULT_INPUT, help="Исходный CELP")
    p.add_argument("--out", type=Path, default=_DEFAULT_OUT, help="Папка usl_spec")
    p.add_argument("--dry-run", action="store_true", help="Только список файлов")
    args = p.parse_args()

    inp = args.input.resolve()
    out = args.out.resolve()
    if str(_REPO_ROOT) not in sys.path:
        sys.path.insert(0, str(_REPO_ROOT))

    run(input_path=inp, out_dir=out, dry_run=args.dry_run)
    if not args.dry_run:
        print(f"Готово: {len(list(out.glob('*.xlsx')))} файлов в {out}")


if __name__ == "__main__":
    main()
