"""
Сборка JSON справочника dash_dn (format: dash_dn_catalog) из Excel в стиле
import_dn_diagnoses_168n / import_dn_services_catalog / import_dn_usl_spec.

Запуск из корня репозитория:
  set PYTHONPATH=%CD%
  py -m apps.dash_dn.build_catalog_from_excel --out apps/dash_dn/data/bundle_global.json

По умолчанию пути к данным — apps/dn_reference/data/...
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable, Optional

# Корень репоз.: .../MosaicMedProject
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
_DEFAULT_DATA = _REPO_ROOT / "apps" / "dn_reference" / "data"

SKIP_PREFIXES = ("~$",)

_MKB_CYRILLIC_TO_LATIN = str.maketrans(
    {
        "А": "A",
        "В": "B",
        "Е": "E",
        "К": "K",
        "М": "M",
        "Н": "H",
        "О": "O",
        "Р": "P",
        "С": "C",
        "Т": "T",
        "У": "Y",
        "Х": "X",
    }
)


def _norm(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if s.lower() in {"nan", "none"}:
        return ""
    return s


def _norm_code(value: object) -> str:
    s = _norm(value).upper().replace(" ", "")
    return s.translate(_MKB_CYRILLIC_TO_LATIN)


def _mkb_norm(value: object) -> str:
    s = _norm(value).upper()
    if not s or s == "-":
        return ""
    return s.split()[0]


def _fix_truncated_c_chapter_range(token: str) -> str:
    m = re.fullmatch(r"(\d{2})-(C[\d.].*)", token, flags=re.IGNORECASE)
    if not m:
        return token
    left, right = m.group(1), m.group(2).upper()
    return f"C{left}-{right}"


def _is_plus(value: object) -> bool:
    s = _norm(value)
    return s in {"+", "＋", "✓", "да", "Да", "YES", "yes", "1"}


def _expand_range_token(token: str) -> set[str]:
    token = token.replace("–", "-").replace("—", "-").replace(" ", "")
    if "-" not in token:
        return {token}

    left, right = token.split("-", 1)
    if not left or not right:
        return {token}

    def split_icd(code: str):
        code = code.upper()
        if not code:
            return None
        letter = code[0]
        rest = code[1:]
        if "." in rest:
            major, minor = rest.split(".", 1)
            if not major.isdigit() or not minor.isdigit():
                return None
            return letter, int(major), int(minor), True
        if not rest.isdigit():
            return None
        return letter, int(rest), None, False

    l = split_icd(left)
    r = split_icd(right)
    if not l or not r:
        return {token}
    if l[0] != r[0]:
        return {token}

    letter = l[0]
    if not l[3] and not r[3]:
        if l[1] > r[1]:
            return {token}
        return {f"{letter}{num:02d}" for num in range(l[1], r[1] + 1)}

    if l[1] == r[1] and l[3] and r[3]:
        if l[2] > r[2]:
            return {token}
        return {f"{letter}{l[1]:02d}.{num}" for num in range(l[2], r[2] + 1)}

    return {token}


def _rule_to_codes(rule: str) -> set[str]:
    codes: set[str] = set()
    for raw in rule.split(","):
        token = _norm_code(raw)
        if not token:
            continue
        token = _fix_truncated_c_chapter_range(token)
        codes.update(_expand_range_token(token))
    return codes


def _diagnosis_matches_rule(diagnosis_code: str, rule_codes: set[str]) -> bool:
    diagnosis_code = _norm_code(diagnosis_code)
    if not diagnosis_code:
        return False
    primary = diagnosis_code.split()[0]
    base = primary.split(".")[0]
    prefixes = {primary, base}
    if "." in primary:
        major, minor = primary.split(".", 1)
        prefixes.add(f"{major}.{minor}")
    return any(
        diagnosis_code.startswith(code) or primary.startswith(code) or base == code
        for code in rule_codes
    )


def _split_joint_specialties(value: str) -> list[str]:
    if not value:
        return []
    parts = [p.strip() for p in value.split("/") if p.strip()]
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        key = p.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


@dataclass(frozen=True)
class Row168n:
    ds: str
    speciality: str
    joint_speciality: str
    category: str


def _iter_rows_168n(path: Path, sheet: Optional[str]) -> Iterable[Row168n]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    header_l = [h.lower() for h in header]

    def col_idx(*names: str) -> int:
        for n in names:
            n_l = n.lower()
            if n_l in header_l:
                return header_l.index(n_l)
        raise ValueError(f"Не найдена колонка {names} в заголовках: {header}")

    i_ds = col_idx("ds", "диагноз", "код", "мкб", "mkb")
    i_spec = col_idx("speciality", "specialty", "специальность")
    try:
        i_joint = col_idx("joint_speciality", "joint_specialty", "совместная специальность", "общая специальность")
    except ValueError:
        i_joint = None
    i_group = col_idx("group", "категория", "группа")

    for row in ws.iter_rows(min_row=2):
        values = [c.value for c in row]
        yield Row168n(
            ds=_mkb_norm(values[i_ds] if i_ds < len(values) else ""),
            speciality=_norm(values[i_spec] if i_spec < len(values) else ""),
            joint_speciality=_norm(values[i_joint] if i_joint is not None and i_joint < len(values) else ""),
            category=_norm(values[i_group] if i_group < len(values) else ""),
        )
    wb.close()


PERIOD_RE = re.compile(
    r"^с\s+(?P<start>\d{2}\.\d{2}\.\d{4})(?:\s+по\s+(?P<end>\d{2}\.\d{2}\.\d{4}))?$",
    re.IGNORECASE,
)


def _parse_date(value: str):
    return datetime.strptime(value, "%d.%m.%Y").date()


def _parse_price(value: object):
    raw = _norm(value).replace(" ", "").replace(",", ".")
    if not raw:
        return None
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _detect_usl_headers(ws) -> tuple[int, int, list[tuple[int, str]]]:
    row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    headers = [_norm(v) for v in row]
    headers_l = [h.lower() for h in headers]

    def idx_by_contains(*variants: str, default: int | None = None):
        for i, h in enumerate(headers_l):
            if all(v.lower() in h for v in variants):
                return i
        return default

    name_idx = idx_by_contains("наименование", default=1)
    code_idx = idx_by_contains("код", default=2)
    if name_idx is None or code_idx is None:
        raise ValueError(f"Не найдены колонки услуги в заголовке: {headers}")

    group_cols: list[tuple[int, str]] = []
    for i, header in enumerate(headers):
        if i in (name_idx, code_idx):
            continue
        if header:
            group_cols.append((i, header))

    if not group_cols:
        raise ValueError(f"Не найдены группы диагнозов в заголовке: {headers}")

    return name_idx, code_idx, group_cols


TABLES_ORDER = [
    "dn_diagnosis_category",
    "dn_specialty",
    "dn_diagnosis_group",
    "dn_service",
    "dn_service_price_period",
    "dn_diagnosis",
    "dn_diagnosis_specialty",
    "dn_diagnosis_group_membership",
    "dn_service_price",
    "dn_service_requirement",
]


class CatalogBuilder:
    def __init__(self, catalog: str = "global") -> None:
        self.catalog = catalog
        self.tables: dict[str, list[dict[str, Any]]] = {t: [] for t in TABLES_ORDER}
        self._seq: dict[str, int] = {t: 0 for t in TABLES_ORDER}

        self._cat_by_name: dict[str, int] = {}
        self._spec_by_name: dict[str, int] = {}
        self._diag_by_code: dict[str, int] = {}
        self._grp_by_code: dict[str, int] = {}
        self._srv_by_code: dict[str, int] = {}
        self._period_by_key: dict[tuple[str, str | None], int] = {}

    def _nid(self, table: str) -> int:
        self._seq[table] += 1
        return self._seq[table]

    def _row(self, table: str, **fields: Any) -> dict[str, Any]:
        rid = self._nid(table)
        r = {"id": rid, "catalog": self.catalog, **fields}
        self.tables[table].append(r)
        return r

    def get_or_create_category(self, name: str) -> int:
        if not name:
            raise ValueError("пустая категория")
        if name in self._cat_by_name:
            return self._cat_by_name[name]
        self._row("dn_diagnosis_category", name=name, is_active=1)
        self._cat_by_name[name] = rid = self.tables["dn_diagnosis_category"][-1]["id"]
        return rid

    def get_or_create_specialty(self, name: str) -> int:
        if not name:
            raise ValueError("пустая специальность")
        if name in self._spec_by_name:
            return self._spec_by_name[name]
        self._row("dn_specialty", name=name, is_active=1)
        self._spec_by_name[name] = rid = self.tables["dn_specialty"][-1]["id"]
        return rid

    def get_or_create_diagnosis(self, code: str, category_id: Optional[int]) -> int:
        if code in self._diag_by_code:
            return self._diag_by_code[code]
        self._row("dn_diagnosis", code=code, category_id=category_id, is_active=1)
        self._diag_by_code[code] = rid = self.tables["dn_diagnosis"][-1]["id"]
        return rid

    def link_diagnosis_specialty(self, diagnosis_id: int, specialty_id: int, source: str) -> None:
        key = (diagnosis_id, specialty_id, source)
        for r in self.tables["dn_diagnosis_specialty"]:
            if (
                r["diagnosis_id"] == diagnosis_id
                and r["specialty_id"] == specialty_id
                and r["source"] == source
            ):
                return
        self._row(
            "dn_diagnosis_specialty",
            diagnosis_id=diagnosis_id,
            specialty_id=specialty_id,
            source=source,
        )

    def get_or_create_group(self, header: str, order: int) -> int:
        code = _norm_code(header)[:256]
        if not code:
            code = f"G{order}"
        if code in self._grp_by_code:
            return self._grp_by_code[code]
        self._row(
            "dn_diagnosis_group",
            code=code,
            title=header[:255],
            sort_order=order,
            rule=header,
            is_active=1,
        )
        self._grp_by_code[code] = rid = self.tables["dn_diagnosis_group"][-1]["id"]
        return rid

    def sync_group_memberships(self, group_id: int, rule: str) -> None:
        rule_codes = _rule_to_codes(rule or "")
        if not rule_codes:
            return
        for code, did in self._diag_by_code.items():
            if _diagnosis_matches_rule(code, rule_codes):
                seen = False
                for r in self.tables["dn_diagnosis_group_membership"]:
                    if r["group_id"] == group_id and r["diagnosis_id"] == did:
                        seen = True
                        break
                if not seen:
                    self._row(
                        "dn_diagnosis_group_membership",
                        group_id=group_id,
                        diagnosis_id=did,
                        is_active=1,
                    )

    def get_or_create_service(self, code: str, name: str, sort_order: int) -> int:
        if code in self._srv_by_code:
            sid = self._srv_by_code[code]
            for r in self.tables["dn_service"]:
                if r["id"] == sid:
                    if name and r.get("name") != name:
                        r["name"] = name
                    break
            return sid
        self._row("dn_service", code=code, name=name or code, sort_order=sort_order, is_active=1)
        self._srv_by_code[code] = self.tables["dn_service"][-1]["id"]
        return self._srv_by_code[code]

    def get_or_create_period(self, date_start: str, date_end: Optional[str], title: str) -> int:
        key = (date_start, date_end)
        if key in self._period_by_key:
            return self._period_by_key[key]
        self._row(
            "dn_service_price_period",
            date_start=date_start,
            date_end=date_end,
            title=title[:128] if title else "",
            is_active=1,
        )
        self._period_by_key[key] = self.tables["dn_service_price_period"][-1]["id"]
        return self._period_by_key[key]

    def add_price(self, service_id: int, period_id: int, price: float) -> None:
        for r in self.tables["dn_service_price"]:
            if r["service_id"] == service_id and r["period_id"] == period_id:
                r["price"] = float(price)
                return
        self._row("dn_service_price", service_id=service_id, period_id=period_id, price=float(price))

    def add_requirement(self, service_id: int, group_id: int, specialty_id: int) -> None:
        for r in self.tables["dn_service_requirement"]:
            if (
                r["service_id"] == service_id
                and r["group_id"] == group_id
                and r["specialty_id"] == specialty_id
            ):
                return
        self._row(
            "dn_service_requirement",
            service_id=service_id,
            group_id=group_id,
            specialty_id=specialty_id,
            is_required=1,
        )

    def to_payload(self) -> dict[str, Any]:
        return {
            "format": "dash_dn_catalog",
            "version": 1,
            "catalog": self.catalog,
            "tables": {k: v for k, v in self.tables.items()},
        }


def load_diagnoses_168n(builder: CatalogBuilder, path: Path, sheet: Optional[str]) -> None:
    for r in _iter_rows_168n(path, sheet):
        if not r.ds:
            continue
        cat_id: Optional[int] = None
        if r.category:
            cat_id = builder.get_or_create_category(r.category)
        did = builder.get_or_create_diagnosis(r.ds, cat_id)
        if r.speciality:
            sid = builder.get_or_create_specialty(r.speciality)
            builder.link_diagnosis_specialty(did, sid, "primary")
        for joint in _split_joint_specialties(r.joint_speciality):
            sid = builder.get_or_create_specialty(joint)
            builder.link_diagnosis_specialty(did, sid, "joint")


def load_services_catalog(builder: CatalogBuilder, path: Path, sheet: Optional[str]) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.worksheets[0]

    header_row_idx = 1
    headers: list[str] = []
    headers_l: list[str] = []

    for r_idx in range(1, 11):
        row = list(ws.iter_rows(min_row=r_idx, max_row=r_idx))[0]
        vals = [_norm(c.value) for c in row]
        vals_l = [v.lower() for v in vals]
        has_num = "№" in " ".join(vals) or "п/п" in " ".join(vals)
        has_name = "наименование" in " ".join(vals_l) and "услуг" in " ".join(vals_l)
        has_code = "код" in " ".join(vals_l) and "услуг" in " ".join(vals_l)
        if has_name and has_code:
            header_row_idx = r_idx
            headers = vals
            headers_l = vals_l
            break

    if not headers:
        row1 = list(ws.iter_rows(min_row=1, max_row=1))[0]
        headers = [_norm(c.value) for c in row1]
        headers_l = [h.lower() for h in headers]

    def col_idx(*variants: str) -> int:
        for v in variants:
            v_l = v.lower()
            for i, h in enumerate(headers_l):
                if v_l in h:
                    return i
        raise ValueError(f"Не найдена колонка {variants} в заголовках: {headers}")

    try:
        i_num = col_idx("№", "п/п", "n", "номер", "no")
    except ValueError:
        i_num = 0
    try:
        i_name = col_idx("наименование", "услуг", "name", "название")
    except ValueError:
        i_name = 1
    try:
        i_code = col_idx("код", "услуг", "code")
    except ValueError:
        i_code = 2

    period_columns: list[tuple[int, tuple]] = []
    for idx, header in enumerate(headers):
        header_norm = _norm(header)
        match = PERIOD_RE.match(header_norm)
        if match:
            start = _parse_date(match.group("start"))
            end = _parse_date(match.group("end")) if match.group("end") else None
            period_columns.append((idx, (start, end, header_norm)))

    order_val = 0
    for row in ws.iter_rows(min_row=header_row_idx + 1):
        values = [c.value for c in row]
        name = _norm(values[i_name] if i_name < len(values) else "")
        code = _norm(values[i_code] if i_code < len(values) else "")
        num_raw = values[i_num] if i_num < len(values) else None
        if not code:
            continue
        if not name:
            name = code
        order_val += 1
        try:
            if num_raw is not None and str(num_raw).strip().isdigit():
                order_val = int(num_raw)
        except (ValueError, TypeError):
            pass

        sid = builder.get_or_create_service(code, name, order_val)

        for cidx, (date_start, date_end, period_title) in period_columns:
            if cidx >= len(values):
                continue
            price = _parse_price(values[cidx])
            if price is None:
                continue
            ds = date_start.isoformat()
            de = date_end.isoformat() if date_end else None
            pid = builder.get_or_create_period(ds, de, period_title)
            builder.add_price(sid, pid, float(price))
    wb.close()


def load_usl_spec_dir(builder: CatalogBuilder, base: Path) -> None:
    import openpyxl

    for path in sorted(base.glob("*.xlsx")):
        if any(path.name.startswith(p) for p in SKIP_PREFIXES):
            continue
        specialty_name = path.stem.strip()
        specialty_id = builder.get_or_create_specialty(specialty_name)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        name_idx, code_idx, group_cols = _detect_usl_headers(ws)
        groups_by_col: dict[int, int] = {}

        for order, (col_idx, group_header) in enumerate(group_cols, start=1):
            gid = builder.get_or_create_group(group_header, order)
            groups_by_col[col_idx] = gid
            builder.sync_group_memberships(gid, group_header)

        for row in ws.iter_rows(min_row=2, values_only=True):
            values = list(row)
            service_name = _norm(values[name_idx] if name_idx < len(values) else "")
            service_code = _norm(values[code_idx] if code_idx < len(values) else "")
            if not service_code:
                continue
            if not service_name:
                service_name = service_code
            srv_rows = builder.tables["dn_service"]
            max_so = max((r.get("sort_order") or 0) for r in srv_rows) if srv_rows else 0
            srv_id = builder.get_or_create_service(service_code, service_name, max_so + 1)

            for col_idx, gid in groups_by_col.items():
                cell = values[col_idx] if col_idx < len(values) else None
                if not _is_plus(cell):
                    continue
                builder.add_requirement(srv_id, gid, specialty_id)
        wb.close()


def main() -> None:
    p = argparse.ArgumentParser(description="Excel → JSON dash_dn_catalog")
    p.add_argument("--out", type=Path, default=_REPO_ROOT / "apps" / "dash_dn" / "data" / "bundle_global.json")
    p.add_argument("--diagnoses", type=Path, default=_DEFAULT_DATA / "диагнозы по 168н.xlsx")
    p.add_argument("--services", type=Path, default=_DEFAULT_DATA / "свод услуг.xlsx")
    p.add_argument("--usl-spec", type=Path, default=_DEFAULT_DATA / "usl_spec")
    p.add_argument("--sheet-168n", type=str, default=None)
    p.add_argument("--sheet-services", type=str, default=None)
    p.add_argument("--also-seed", action="store_true", help="Перезаписать apps/dash_dn/data/seed_global.json")
    p.add_argument(
        "--apply-sqlite",
        action="store_true",
        help="Импортировать результат в SQLite dash_dn (DASH_DN_SQLITE или data/dn_catalog.sqlite), каталог global",
    )
    args = p.parse_args()

    root = _REPO_ROOT
    if str(root) not in sys.path:
        sys.path.insert(0, str(root))

    if not args.diagnoses.is_file():
        print(f"Пропуск 168н (нет файла): {args.diagnoses}", file=sys.stderr)
    if not args.services.is_file():
        print(f"Пропуск свода услуг (нет файла): {args.services}", file=sys.stderr)
    if not args.usl_spec.is_dir():
        print(f"Пропуск usl_spec (нет папки): {args.usl_spec}", file=sys.stderr)

    builder = CatalogBuilder("global")

    if args.diagnoses.is_file():
        print("Загрузка диагнозов 168н…")
        load_diagnoses_168n(builder, args.diagnoses, args.sheet_168n)
    if args.services.is_file():
        print("Загрузка свода услуг…")
        load_services_catalog(builder, args.services, args.sheet_services)
    if args.usl_spec.is_dir():
        print("Загрузка usl_spec…")
        load_usl_spec_dir(builder, args.usl_spec)

    payload = builder.to_payload()
    args.out.parent.mkdir(parents=True, exist_ok=True)
    with args.out.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"Записано: {args.out} (таблицы: { {k: len(v) for k, v in payload['tables'].items()} })")

    if args.also_seed:
        seed_path = _REPO_ROOT / "apps" / "dash_dn" / "data" / "seed_global.json"
        with seed_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"Обновлён seed: {seed_path}")

    if args.apply_sqlite:
        from apps.dash_dn.sqlite_catalog.catalog_ops import import_catalog, load_json_file
        from apps.dash_dn.sqlite_catalog.db import ensure_database
        from apps.dash_dn.sqlite_catalog.paths import get_db_path

        eng = ensure_database()
        data = load_json_file(args.out)
        with eng.begin() as conn:
            import_catalog(conn, data, target_catalog="global")
        print(f"SQLite обновлена (global): {get_db_path()}")


if __name__ == "__main__":
    main()
