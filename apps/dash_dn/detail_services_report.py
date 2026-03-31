# -*- coding: utf-8 -*-
"""
Отчёт по выгрузке detail_services_*.csv (свод по диагнозам, талоны 1/2+ услуг,
проверка кодов по справочнику услуг).

По умолчанию допустимые коды берутся из SQLite (таблица dn_service, каталог global).
Опционально — из Excel «свод услуг.xlsx» (как раньше).

Используется CLI `scripts/дн услуги/report_detail_services.py` и вкладка «Анализ» в dash_dn.
"""
from __future__ import annotations

import csv
import io
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from sqlalchemy import bindparam, text
from sqlalchemy.engine import Engine

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
except ImportError as e:  # pragma: no cover
    raise ImportError(
        "Нужен пакет openpyxl: pip install openpyxl"
    ) from e

_PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_REFERENCE = _PROJECT_ROOT / "scripts" / "дн услуги" / "свод услуг.xlsx"

# Базовый фрагмент Axx.xx.xxx / Bxx.xx.xxx + опционально уточнения исследования (.xxx)…
# В справочнике встречаются, например, A06.09.007.002 при том, что валидируется «база» A06.09.007.
SERVICE_CODE_RE = re.compile(
    r"^([AB]\d{2}\.\d{2,3}\.\d{3})(\.\d{3})*$",
    re.IGNORECASE,
)


def _service_code_accepted_by_reference(code: str, allowed_codes: set[str]) -> bool:
    """Код допустим, если есть в справочнике полностью или совпадает база (без хвоста .xxx)."""
    c = _norm_code(code)
    if not c:
        return False
    if c.upper() in allowed_codes:
        return True
    m = SERVICE_CODE_RE.match(c)
    if not m:
        return False
    base = m.group(1).upper()
    return base in allowed_codes


def filter_detail_rows_by_valid_service(
    rows: list[dict[str, str]],
    allowed_codes: set[str],
    *,
    col_code: str = "Код услуги",
) -> tuple[list[dict[str, str]], dict[str, int]]:
    """Оставляет только строки с кодом услуги из справочника.

    Пропускаются: пусто/прочерк/«нет кода», неверный формат, код не в dn_service
    (включая вариант «база в справочнике» для уточняющего хвоста .xxx).
    """
    kept: list[dict[str, str]] = []
    skipped_missing = 0
    skipped_rejected = 0
    for row in rows:
        code = _norm_code(row.get(col_code))
        if _is_missing_code(code):
            skipped_missing += 1
            continue
        if not SERVICE_CODE_RE.match(code):
            skipped_rejected += 1
            continue
        if not _service_code_accepted_by_reference(code, allowed_codes):
            skipped_rejected += 1
            continue
        kept.append(row)
    return kept, {
        "skipped_missing": skipped_missing,
        "skipped_rejected": skipped_rejected,
    }


# МКБ-10 в тексте выгрузки (DS1/DS2): буква + 2 цифры + необязательный хвост .N…
# Первая буква кода МКБ-10 (включая главу U, напр. U07.1)
ICD10_CODE_RE = re.compile(r"\b([A-Z]\d{2}(?:\.\d+)?)\b", re.IGNORECASE)


def resolve_reference_path() -> Path:
    env = os.environ.get("DASH_DN_DETAIL_SERVICES_REF")
    if env:
        return Path(env.strip()).resolve()
    return DEFAULT_REFERENCE.resolve()


def _norm_code(s: str | None) -> str:
    if s is None:
        return ""
    t = str(s).strip().replace("\u00a0", " ")
    return t


# Нормализованное значение ячейки «Код услуги» (без пробелов, нижний регистр) —
# пометки «нет услуги/кода» как у Квазар; такие строки не сверяем со справочником.
_SERVICE_CODE_MISSING_PHRASES = frozenset(
    {
        "безкода",
        "безуслуги",
        "неткода",
        "нетуслуги",
        "отсутствует",
        "н/д",
        "н-д",
        "n/a",
        "na",
        "none",
        "-",
        "—",
        ".",
        "–",
    }
)


def _is_missing_code(code: str) -> bool:
    """Пусто, прочерк или явная пометка «без услуги/кода» в выгрузке — строка не проверяется по справочнику."""
    if not code:
        return True
    t = _norm_code(code)
    if not t:
        return True
    if t in ("-", "—", ".", "–"):
        return True
    key = t.casefold().replace(" ", "").replace("\t", "")
    if key in _SERVICE_CODE_MISSING_PHRASES:
        return True
    return False


def _parse_money(s: str | None) -> float:
    if s is None or str(s).strip() in ("", "-", "—"):
        return 0.0
    t = str(s).strip().replace("\u00a0", " ").replace(" ", "").replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def load_allowed_service_codes(xlsx_path: Path) -> set[str]:
    """Читает коды услуг из 3-й колонки первого листа свода (колонка «Код»)."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    allowed: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        if len(row) < 3:
            continue
        raw = row[2]
        if raw is None:
            continue
        c = _norm_code(str(raw))
        if not c or _is_missing_code(c):
            continue
        allowed.add(c.upper())
    wb.close()
    return allowed


def load_diagnosis_lookup_from_db(
    catalog: str = "global",
    *,
    engine: Engine | None = None,
) -> dict[str, str | None]:
    """Код МКБ (верхний регистр) → название категории из dn_diagnosis_category или None (диагноз есть, категория пустая)."""
    from apps.dash_dn.sqlite_catalog.db import get_engine

    eng = engine or get_engine()
    q = text(
        """
        SELECT UPPER(TRIM(d.code)) AS code, c.name AS category_name
        FROM dn_diagnosis d
        LEFT JOIN dn_diagnosis_category c
          ON c.id = d.category_id AND c.catalog = :cat
        WHERE d.catalog = :cat AND d.is_active = 1
        """
    )
    out: dict[str, str | None] = {}
    with eng.connect() as conn:
        for row in conn.execute(q, {"cat": catalog}).fetchall():
            code, cat = row[0], row[1]
            if code:
                out[code] = cat if cat else None
    return out


def load_required_services_for_icd_codes(
    icd_codes: set[str] | list[str],
    catalog: str = "global",
    *,
    engine: Engine | None = None,
) -> dict[str, list[tuple[str, str]]]:
    """Код МКБ (верхний регистр) → список (код услуги, название) из dn_service_requirement
    через группы диагнозов (dn_diagnosis_group_membership). Пустой список — нет связи в справочнике.
    """
    from apps.dash_dn.sqlite_catalog.db import get_engine

    codes = sorted({c.upper().strip() for c in icd_codes if c and str(c).strip()})
    out: dict[str, list[tuple[str, str]]] = {c: [] for c in codes}
    if not codes:
        return {}
    eng = engine or get_engine()
    q = text(
        """
        SELECT DISTINCT UPPER(TRIM(d.code)) AS dcode, srv.code AS scode, srv.name AS sname
        FROM dn_diagnosis d
        JOIN dn_diagnosis_group_membership gm
          ON gm.diagnosis_id = d.id AND gm.catalog = :cat AND gm.is_active = 1
        JOIN dn_service_requirement sr
          ON sr.group_id = gm.group_id AND sr.catalog = :cat AND sr.is_required = 1
        JOIN dn_service srv ON srv.id = sr.service_id AND srv.catalog = :cat
        WHERE d.catalog = :cat AND d.is_active = 1
          AND UPPER(TRIM(d.code)) IN :codes
          AND srv.is_active = 1
        ORDER BY dcode, scode
        """
    ).bindparams(bindparam("codes", expanding=True))
    with eng.connect() as conn:
        rows = conn.execute(q, {"cat": catalog, "codes": codes}).fetchall()
    seen: dict[str, set[tuple[str, str]]] = defaultdict(set)
    for dcode, scode, sname in rows:
        if dcode is None or scode is None:
            continue
        key = str(dcode).upper()
        if key not in out:
            continue
        t = (str(scode), str(sname) if sname is not None else "")
        if t in seen[key]:
            continue
        seen[key].add(t)
        out[key].append(t)
    return out


def _service_codes_only(items: list[tuple[str, str]]) -> list[str]:
    """Только коды услуг, без названий; порядок как в справочнике, без дубликатов."""
    out: list[str] = []
    seen: set[str] = set()
    for sc, _ in items:
        if sc not in seen:
            seen.add(sc)
            out.append(sc)
    return out


def _service_code_base(code: str | None) -> str:
    """База кода услуги для сопоставления с обязательными (Axx.xx.xxx)."""
    c = _norm_code(code)
    if not c:
        return ""
    m = SERVICE_CODE_RE.match(c)
    if m:
        return m.group(1).upper()
    return c.upper()


def _required_service_codes_for_ticket(
    ds1_code: str,
    ds2_codes: list[str],
    svc_by_icd: dict[str, list[tuple[str, str]]],
) -> list[str]:
    """Обязательные коды услуг по DS1/DS2 без дублей, в порядке: DS1, затем DS2."""
    out: list[str] = []
    seen: set[str] = set()
    seq: list[str] = []
    if ds1_code:
        seq.append(ds1_code)
    seq.extend(ds2_codes)
    for icd in seq:
        for sc, _ in svc_by_icd.get(icd.upper(), []):
            base = _service_code_base(sc)
            if not base or base in seen:
                continue
            seen.add(base)
            out.append(base)
    return out


def _present_service_code_bases(
    trows: list[dict[str, str]],
    allowed_codes: set[str],
) -> set[str]:
    """Коды услуг талона (база Axx.xx.xxx), только валидные и из справочника услуг."""
    out: set[str] = set()
    for r in trows:
        code = _norm_code(r.get("Код услуги"))
        if _is_missing_code(code):
            continue
        if not SERVICE_CODE_RE.match(code):
            continue
        if not _service_code_accepted_by_reference(code, allowed_codes):
            continue
        out.add(_service_code_base(code))
    return out


def _ticket_required_services_issues(
    ds1_code: str,
    ds2_codes: list[str],
    trows: list[dict[str, str]],
    allowed_codes: set[str],
    svc_by_icd: dict[str, list[tuple[str, str]]],
) -> list[str]:
    """Проверка обязательных услуг в мягком режиме:
    если обязательные есть, должна быть хотя бы одна из них; лишние не считаются ошибкой.
    """
    # Для текущего свода ориентируемся на основной диагноз (DS1):
    # это убирает ложные срабатывания из-за сопутствующих DS2.
    required = _required_service_codes_for_ticket(ds1_code, [], svc_by_icd)
    if not required:
        return []
    present = _present_service_code_bases(trows, allowed_codes)
    matched = [c for c in required if c in present]
    if matched:
        return []
    preview = "; ".join(required[:5])
    if len(required) > 5:
        preview += "; ..."
    return [
        "нет ни одной обязательной услуги по диагнозам (пример обязательных: " + preview + ")",
    ]


def _ticket_services_suggestion_text(
    ds1_code: str,
    ds2_codes: list[str],
    svc_by_icd: dict[str, list[tuple[str, str]]],
) -> str:
    """Текст для отчёта: обязательные услуги по основному и сопутствующим МКБ из справочника ДН.
    Коды услуг не повторяются между основным и сопутствующими (и между сопутствующими): уже
    перечисленные выше не дублируются в следующих блоках.
    """
    parts: list[str] = []
    already_listed: set[str] = set()

    if ds1_code:
        u = ds1_code.upper()
        items = svc_by_icd.get(u, [])
        lab = f"Основной ({ds1_code})"
        if not items:
            parts.append(
                f"{lab}: нет обязательных услуг в справочнике "
                f"(группа dn_diagnosis_group / dn_service_requirement)"
            )
        else:
            codes = _service_codes_only(items)
            for x in codes:
                already_listed.add(x)
            parts.append(lab + ": " + "; ".join(codes))
    for c in ds2_codes:
        u = c.upper()
        items = svc_by_icd.get(u, [])
        lab = f"Сопутствующий ({c})"
        if not items:
            parts.append(
                f"{lab}: нет обязательных услуг в справочнике "
                f"(группа dn_diagnosis_group / dn_service_requirement)"
            )
        else:
            raw = _service_codes_only(items)
            new_only = [x for x in raw if x not in already_listed]
            if not new_only:
                continue
            for x in new_only:
                already_listed.add(x)
            parts.append(lab + ": " + "; ".join(new_only))
    return " | ".join(parts) if parts else ""


def _icd_first_code(cell: str | None) -> str:
    if not cell:
        return ""
    m = ICD10_CODE_RE.search(_norm_code(cell))
    return m.group(1).upper() if m else ""


def _icd_codes_in_text(s: str) -> list[str]:
    """Все уникальные коды МКБ по порядку появления в строке (DS2 может содержать несколько)."""
    if not s or _is_missing_code(s.strip()):
        return []
    seen: set[str] = set()
    out: list[str] = []
    for m in ICD10_CODE_RE.finditer(s):
        c = m.group(1).upper()
        if c not in seen:
            seen.add(c)
            out.append(c)
    return out


def _category_label_for_icd(
    code: str,
    lookup: dict[str, str | None],
) -> tuple[str, bool]:
    """(подпись для отчёта, код есть в dn_diagnosis)."""
    if not code:
        return "", True
    u = code.upper()
    if u not in lookup:
        return "нет в справочнике (dn_diagnosis)", False
    cat = lookup[u]
    if cat is None or cat == "":
        return "(без категории в справочнике)", True
    return cat, True


def load_allowed_service_codes_from_db(
    catalog: str = "global",
    *,
    engine: Engine | None = None,
) -> set[str]:
    """Коды активных услуг из локальной SQLite (dn_service), как во вкладке «Подбор услуг ДН»."""
    from apps.dash_dn.sqlite_catalog.db import get_engine

    eng = engine or get_engine()
    q = text(
        """
        SELECT UPPER(TRIM(code)) AS c
        FROM dn_service
        WHERE catalog = :cat AND is_active = 1
          AND code IS NOT NULL AND TRIM(code) != ''
        """
    )
    with eng.connect() as conn:
        rows = conn.execute(q, {"cat": catalog}).fetchall()
    return {r[0] for r in rows if r[0]}


def pick_csv_path(script_dir: Path, explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit.resolve()
    d = script_dir / "выгрузки"
    candidates = sorted(
        d.glob("detail_services_*.csv"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(
            f"Не найден CSV в {d}. Укажите --csv путь к файлу."
        )
    return candidates[0]


def read_detail_csv_from_text(text: str) -> tuple[list[str], list[dict[str, str]]]:
    f = io.StringIO(text)
    r = csv.DictReader(f, delimiter=";")
    if not r.fieldnames:
        raise ValueError("Пустой CSV или нет заголовка")
    fieldnames = list(r.fieldnames)
    rows = [dict(row) for row in r]
    return fieldnames, rows


def read_detail_csv_from_bytes(raw: bytes) -> tuple[list[str], list[dict[str, str]]]:
    text = raw.decode("utf-8-sig")
    return read_detail_csv_from_text(text)


def read_detail_csv(csv_path: Path) -> tuple[list[str], list[dict[str, str]]]:
    return read_detail_csv_from_bytes(csv_path.read_bytes())


def _status_from_row(row: dict[str, str]) -> str:
    v = row.get("Статус")
    if v is None:
        return ""
    return _norm_code(str(v))


def _format_services_for_ticket(trows: list[dict[str, str]]) -> str:
    """Строка для отчёта: «код — сумма» по каждой строке талона, через «; ».

    Если в «Код услуги» прочерк/пусто — пометка «без кода в выгрузке».
    """
    parts: list[str] = []
    for r in trows:
        code = _norm_code(r.get("Код услуги"))
        if _is_missing_code(code):
            label = "без кода в выгрузке"
        else:
            label = code
        sm = _parse_money(r.get("Сумма услуги"))
        qty_raw = r.get("Кол-во услуг")
        try:
            if qty_raw in (None, "", "-", "—"):
                qty = 1.0
            else:
                qty = float(str(qty_raw).strip().replace(",", ".").replace(" ", ""))
        except ValueError:
            qty = 1.0
        if qty != 1.0:
            parts.append(f"{label} — {sm} (×{qty:g})")
        else:
            parts.append(f"{label} — {sm}")
    return "; ".join(parts)


def _service_line_issues(
    trows: list[dict[str, str]],
    allowed_codes: set[str],
) -> list[str]:
    """Причины по строкам талона (нумерация с 1 — порядок строк в выгрузке по талону)."""
    col_code = "Код услуги"
    issues: list[str] = []
    for i, r in enumerate(trows, start=1):
        code = _norm_code(r.get(col_code))
        if _is_missing_code(code):
            continue
        if not SERVICE_CODE_RE.match(code):
            issues.append(
                f"услуга талона (стр. {i}): «{code}» — неверный формат кода "
                f"(нужен Axx.xx.xxx / Bxx.xx.xxx, при необходимости с уточнением .xxx)"
            )
            continue
        if not _service_code_accepted_by_reference(code, allowed_codes):
            issues.append(
                f"услуга талона (стр. {i}): «{code}» — нет в справочнике услуг"
            )
    return issues


def _ticket_service_lines_match_reference(
    trows: list[dict[str, str]],
    allowed_codes: set[str],
) -> bool:
    """Все непустые строки услуг — допустимый формат и код в справочнике."""
    return not _service_line_issues(trows, allowed_codes)


def _main_category_bucket(
    ds1_code: str,
    ds1_display: str,
    lookup: dict[str, str | None],
) -> str:
    """Ключ для свода по категориям основного диагноза."""
    if not ds1_code:
        if ds1_display and ds1_display != "(без диагноза)":
            return "(нет извлечённого кода МКБ в DS1)"
        return "(нет диагноза DS1)"
    lab, ok = _category_label_for_icd(ds1_code, lookup)
    if not ok:
        return "нет в справочнике (основной)"
    return lab


def build_report(
    rows: list[dict[str, str]],
    allowed_codes: set[str],
    diagnosis_lookup: dict[str, str | None] | None = None,
) -> dict[str, Any]:
    col_ticket = "Талон"
    col_ds1 = "Диагноз основной (DS1)"
    col_ds2 = "Сопутствующий диагноз (DS2)"
    col_code = "Код услуги"
    col_ticket_sum = "Сумма талона"
    col_doctor = "Врач"
    col_unit = "Подразделение"
    col_enp = "ЕНП"

    if not rows:
        raise ValueError("В CSV нет строк данных.")

    for c in (col_ticket, col_ds1, col_code, col_ticket_sum):
        if c not in rows[0]:
            raise KeyError(
                f"В CSV нет колонки «{c}». Проверьте формат выгрузки."
            )
    has_ds2 = col_ds2 in rows[0]
    has_doctor_col = col_doctor in rows[0]
    has_unit_col = col_unit in rows[0]
    has_enp_col = col_enp in rows[0]

    rows_raw_total = len(rows)
    tids_in_file = {
        _norm_code(r.get(col_ticket))
        for r in rows
        if _norm_code(r.get(col_ticket))
    }
    unique_tickets_raw = len(tids_in_file)

    # Этап 1: убрать строки без кода услуги (прочерк «-», пусто, «без кода» и т.п. — см. _is_missing_code).
    skipped_stage1_no_code = sum(
        1 for r in rows if _is_missing_code(_norm_code(r.get(col_code)))
    )
    rows = [
        r
        for r in rows
        if not _is_missing_code(_norm_code(r.get(col_code)))
    ]
    if not rows:
        raise ValueError(
            "После исключения строк без кода услуги (в т.ч. «-») в выгрузке не осталось ни одной строки."
        )

    skip_rejected = 0
    for row in rows:
        code = _norm_code(row.get(col_code))
        if not SERVICE_CODE_RE.match(code):
            skip_rejected += 1
        elif not _service_code_accepted_by_reference(code, allowed_codes):
            skip_rejected += 1
    skip_svc = {
        "skipped_missing": skipped_stage1_no_code,
        "skipped_rejected": skip_rejected,
    }

    ticket_rows: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        tid = _norm_code(row.get(col_ticket))
        if not tid:
            continue
        ticket_rows[tid].append(row)

    tids_kept = set(ticket_rows.keys())
    tickets_dropped_ids = sorted(tids_in_file - tids_kept)
    tickets_dropped_all_lines = len(tickets_dropped_ids)

    all_icd_in_file: set[str] = set()
    for row in rows:
        c1 = _icd_first_code(row.get(col_ds1))
        if c1:
            all_icd_in_file.add(c1.upper())
        if has_ds2:
            for c in _icd_codes_in_text(_norm_code(row.get(col_ds2, ""))):
                all_icd_in_file.add(c.upper())
    svc_by_icd = load_required_services_for_icd_codes(all_icd_in_file, catalog="global")

    dx_lookup = diagnosis_lookup if diagnosis_lookup is not None else {}
    use_dx = diagnosis_lookup is not None

    by_category_ds1: dict[str, dict[str, float | int]] = defaultdict(
        lambda: {
            "талонов": 0,
            "сумма_талонов": 0.0,
            "талонов_услуги_ок": 0,
            "сумма_услуги_ок": 0.0,
        }
    )
    by_category_ds2: dict[str, int] = defaultdict(int)
    by_category_ds1_by_unit: dict[tuple[str, str], dict[str, float | int]] = defaultdict(
        lambda: {
            "талонов": 0,
            "сумма_талонов": 0.0,
            "талонов_услуги_ок": 0,
            "сумма_услуги_ок": 0.0,
        }
    )
    diagnosis_mismatches: list[dict[str, Any]] = []
    ticket_service_issues: dict[str, list[str]] = {}

    per_ticket: list[dict[str, Any]] = []
    for tid, trows in ticket_rows.items():
        first = trows[0]
        ds1_raw = _norm_code(first.get(col_ds1))
        ds1_plain = ds1_raw or ""
        diagnosis = ds1_raw or "(без диагноза)"
        ds2_raw = _norm_code(first.get(col_ds2, "")) if has_ds2 else ""

        ds1_code = _icd_first_code(first.get(col_ds1))
        codes_ds2 = _icd_codes_in_text(ds2_raw)

        n_services = len(trows)
        ticket_sum = _parse_money(first.get(col_ticket_sum))
        doctor_t = _norm_code(first.get(col_doctor)) if has_doctor_col else ""
        unit_t = _norm_code(first.get(col_unit)) if has_unit_col else ""
        enp_t = _norm_code(first.get(col_enp)) if has_enp_col else ""

        ds1_category = ""
        ds2_cat_line = ""
        ds1_ok = True
        svc_issues = _service_line_issues(trows, allowed_codes)
        svc_issues += _ticket_required_services_issues(
            ds1_code,
            codes_ds2,
            trows,
            allowed_codes,
            svc_by_icd,
        )
        ticket_service_issues[tid] = svc_issues
        svc_ok = not svc_issues
        if use_dx:
            lab1, ds1_ok = _category_label_for_icd(ds1_code, dx_lookup)
            ds1_category = lab1 if ds1_code else ""
            parts2: list[str] = []
            for c in codes_ds2:
                lab, ok = _category_label_for_icd(c, dx_lookup)
                parts2.append(f"{c} — {lab}")
                key2 = lab if ok else "нет в справочнике (сопутствующий)"
                by_category_ds2[key2] += 1
            ds2_cat_line = "; ".join(parts2)

            bucket = _main_category_bucket(ds1_code, diagnosis, dx_lookup)
            unit_key = unit_t if unit_t else "(не указано)"
            by_category_ds1[bucket]["талонов"] += 1
            by_category_ds1[bucket]["сумма_талонов"] += float(ticket_sum)
            if svc_ok:
                by_category_ds1[bucket]["талонов_услуги_ок"] += 1
                by_category_ds1[bucket]["сумма_услуги_ок"] += float(ticket_sum)

            cu = (bucket, unit_key)
            by_category_ds1_by_unit[cu]["талонов"] += 1
            by_category_ds1_by_unit[cu]["сумма_талонов"] += float(ticket_sum)
            if svc_ok:
                by_category_ds1_by_unit[cu]["талонов_услуги_ок"] += 1
                by_category_ds1_by_unit[cu]["сумма_услуги_ок"] += float(ticket_sum)

            st = _status_from_row(first)
            dx_ctx = {
                "енп": enp_t,
                "врач": doctor_t,
                "подразделение": unit_t,
            }
            if ds1_plain and not _is_missing_code(ds1_plain) and not ds1_code:
                diagnosis_mismatches.append(
                    {
                        "талон": tid,
                        "статус": st,
                        "тип": "основной",
                        "код_мкб": "",
                        "комментарий": "в DS1 есть текст, код МКБ-10 не извлечён",
                        **dx_ctx,
                    }
                )
            elif ds1_code and not ds1_ok:
                diagnosis_mismatches.append(
                    {
                        "талон": tid,
                        "статус": st,
                        "тип": "основной",
                        "код_мкб": ds1_code,
                        "комментарий": "нет в dn_diagnosis (catalog=global)",
                        **dx_ctx,
                    }
                )
            for c in codes_ds2:
                _, ok = _category_label_for_icd(c, dx_lookup)
                if not ok:
                    diagnosis_mismatches.append(
                        {
                            "талон": tid,
                            "статус": st,
                            "тип": "сопутствующий",
                            "код_мкб": c,
                            "комментарий": "нет в dn_diagnosis (catalog=global)",
                            **dx_ctx,
                        }
                    )
            if ds2_raw and not _is_missing_code(ds2_raw.strip()) and not codes_ds2:
                diagnosis_mismatches.append(
                    {
                        "талон": tid,
                        "статус": st,
                        "тип": "сопутствующий",
                        "код_мкб": "",
                        "комментарий": "в DS2 есть текст, код МКБ-10 не извлечён",
                        **dx_ctx,
                    }
                )
        else:
            ds2_cat_line = ", ".join(codes_ds2) if codes_ds2 else ""

        notes: list[str] = []
        if use_dx:
            if ds1_plain and not _is_missing_code(ds1_plain) and not ds1_code:
                notes.append("DS1: нет кода МКБ")
            elif ds1_code and not ds1_ok:
                notes.append("DS1: нет в справочнике")
            for c in codes_ds2:
                _, ok = _category_label_for_icd(c, dx_lookup)
                if not ok:
                    notes.append(f"{c}: нет в справочнике")
            if ds2_raw and not _is_missing_code(ds2_raw.strip()) and not codes_ds2:
                notes.append("DS2: нет кода МКБ")
        diagnosis_note = "; ".join(notes)

        per_ticket.append(
            {
                "талон": tid,
                "енп": enp_t,
                "статус": _status_from_row(first),
                "врач": doctor_t,
                "подразделение": unit_t,
                "диагноз": diagnosis,
                "ds1_code": ds1_code,
                "ds1_category": ds1_category,
                "ds2_raw": ds2_raw,
                "ds2_codes": ", ".join(codes_ds2),
                "ds2_categories_line": ds2_cat_line,
                "diagnosis_note": diagnosis_note,
                "число_строк_услуг": n_services,
                "сумма_талона": ticket_sum,
                "услуги": _format_services_for_ticket(trows),
                "svc_ok": svc_ok,
            }
        )

    by_diagnosis: dict[str, dict[str, float | int]] = defaultdict(
        lambda: {"талонов": 0, "сумма_талонов": 0.0}
    )
    for t in per_ticket:
        d = str(t["диагноз"])
        by_diagnosis[d]["талонов"] += 1
        by_diagnosis[d]["сумма_талонов"] += float(t["сумма_талона"])

    single_service = [t for t in per_ticket if t["число_строк_услуг"] == 1]
    multi_service = [t for t in per_ticket if t["число_строк_услуг"] > 1]

    invalid_lines: list[dict[str, Any]] = []
    invalid_by_code: dict[str, int] = defaultdict(int)
    no_code_lines: list[dict[str, Any]] = []

    for row in rows:
        tid = _norm_code(row.get(col_ticket))
        code = _norm_code(row.get(col_code))
        ds1 = _norm_code(row.get(col_ds1)) or "(без диагноза)"
        ds2_line = _norm_code(row.get(col_ds2, "")) if has_ds2 else ""
        base_inv = {
            "талон": tid,
            "енп": _norm_code(row.get(col_enp)) if has_enp_col else "",
            "статус": _status_from_row(row),
            "диагноз": ds1,
            "диагноз_ds2": ds2_line,
            "врач": _norm_code(row.get(col_doctor)) if has_doctor_col else "",
            "подразделение": _norm_code(row.get(col_unit)) if has_unit_col else "",
        }
        if _is_missing_code(code):
            no_code_lines.append(
                {
                    **base_inv,
                    "код_услуги": code or "",
                    "сумма_услуги": _parse_money(row.get("Сумма услуги")),
                }
            )
            continue
        if not SERVICE_CODE_RE.match(code):
            invalid_lines.append(
                {
                    **base_inv,
                    "код_услуги": code,
                    "причина": "неверный формат кода",
                    "сумма_услуги": _parse_money(row.get("Сумма услуги")),
                }
            )
            invalid_by_code[code.upper()] += 1
            continue
        if not _service_code_accepted_by_reference(code, allowed_codes):
            invalid_lines.append(
                {
                    **base_inv,
                    "код_услуги": code,
                    "причина": "нет в справочнике услуг",
                    "сумма_услуги": _parse_money(row.get("Сумма услуги")),
                }
            )
            invalid_by_code[code.upper()] += 1

    tickets_service_mismatch: list[dict[str, Any]] = []
    for tid, trows in sorted(ticket_rows.items(), key=lambda x: x[0]):
        issues = ticket_service_issues.get(tid) or []
        if not issues:
            continue
        first = trows[0]
        ds1_c = _icd_first_code(first.get(col_ds1))
        ds2_list = (
            _icd_codes_in_text(_norm_code(first.get(col_ds2, ""))) if has_ds2 else []
        )
        svc_suggest = _ticket_services_suggestion_text(ds1_c, ds2_list, svc_by_icd)
        ds1_mkb = ds1_c.upper() if ds1_c else ""
        ds2_mkb_line = ", ".join(ds2_list) if ds2_list else ""
        tickets_service_mismatch.append(
            {
                "талон": tid,
                "енп": _norm_code(first.get(col_enp)) if has_enp_col else "",
                "статус": _status_from_row(first),
                "врач": _norm_code(first.get(col_doctor)) if has_doctor_col else "",
                "подразделение": _norm_code(first.get(col_unit)) if has_unit_col else "",
                "диагноз": ds1_mkb,
                "сопутствующие_мкб": ds2_mkb_line,
                "услуги_по_диагнозам": svc_suggest,
                "причины": "; ".join(issues),
                "число_строк_услуг": len(trows),
                "сумма_талона": _parse_money(first.get(col_ticket_sum)),
            }
        )

    by_cat_unit_list: list[dict[str, Any]] = []
    if use_dx and by_category_ds1_by_unit:
        for (cat, unit), agg in sorted(
            by_category_ds1_by_unit.items(),
            key=lambda x: (-x[1]["талонов"], x[0][0], x[0][1]),
        ):
            by_cat_unit_list.append(
                {
                    "категория": cat,
                    "подразделение": unit,
                    "талонов": agg["талонов"],
                    "сумма_талонов": round(float(agg["сумма_талонов"]), 2),
                    "талонов_услуги_ок": int(agg["талонов_услуги_ок"]),
                    "сумма_услуги_ок": round(float(agg["сумма_услуги_ок"]), 2),
                }
            )

    return {
        "per_ticket": per_ticket,
        "by_diagnosis": by_diagnosis,
        "by_category_ds1": dict(by_category_ds1) if use_dx else {},
        "by_category_ds1_by_unit": by_cat_unit_list if use_dx else [],
        "by_category_ds2": dict(sorted(by_category_ds2.items(), key=lambda x: (-x[1], x[0]))),
        "diagnosis_mismatches": diagnosis_mismatches,
        "diagnosis_enriched": use_dx,
        "single_service": single_service,
        "multi_service": multi_service,
        "invalid_lines": invalid_lines,
        "invalid_by_code": dict(sorted(invalid_by_code.items(), key=lambda x: -x[1])),
        "no_code_lines": no_code_lines,
        "unique_tickets": len(ticket_rows),
        "unique_tickets_raw": unique_tickets_raw,
        "tickets_dropped_all_lines": tickets_dropped_all_lines,
        "tickets_dropped_ids": tickets_dropped_ids,
        "total_rows": len(rows),
        "rows_raw_total": rows_raw_total,
        "skipped_service_missing": skip_svc["skipped_missing"],
        "skipped_service_rejected": skip_svc["skipped_rejected"],
        "has_ds2_column": has_ds2,
        "has_doctor_column": has_doctor_col,
        "has_unit_column": has_unit_col,
        "has_enp_column": has_enp_col,
        "tickets_service_mismatch": tickets_service_mismatch,
    }


def write_xlsx(
    out_path: Path,
    report: dict[str, Any],
    *,
    source_csv: Path,
    reference_note: str,
) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws0 = wb.active
    ws0.title = "Параметры"
    ws0.append(["Источник CSV", str(source_csv)])
    ws0.append(["Справочник услуг", reference_note])
    ws0.append(["Строк в файле", report.get("rows_raw_total", report["total_rows"])])
    ws0.append(["Уникальных талонов", report["unique_tickets"]])
    ws0.append(
        [
            "Строк услуг: поле «код» пусто или прочерк",
            report.get("skipped_service_missing", 0),
        ]
    )
    ws0.append(
        [
            "Строк услуг: код не в справочнике или неверный формат",
            report.get("skipped_service_rejected", 0),
        ]
    )
    ws0.append(
        [
            "Талонов с замечаниями по услугам",
            len(report.get("tickets_service_mismatch") or []),
        ]
    )
    if report.get("diagnosis_enriched"):
        ws0.append(
            [
                "Справочник диагнозов",
                "dn_diagnosis + dn_diagnosis_category (catalog=global)",
            ]
        )
        ws0.append(
            [
                "Записей о несоответствии диагнозов",
                len(report.get("diagnosis_mismatches") or []),
            ]
        )

    dropped_ids = report.get("tickets_dropped_ids") or []
    if dropped_ids:
        ws_d = wb.create_sheet("Талоны без валидной услуги")
        ws_d.append(
            [
                "Талон",
                "Комментарий",
            ]
        )
        note = "Все строки «Код услуги» — пусто/прочерк или код не в dn_service"
        for tid in dropped_ids:
            ws_d.append([tid, note])

    ws1 = wb.create_sheet("Свод по диагнозам")
    ws1.append(["Диагноз основной (DS1)", "Талонов", "Сумма талонов"])
    for diag, agg in sorted(
        report["by_diagnosis"].items(),
        key=lambda x: (-x[1]["талонов"], x[0]),
    ):
        ws1.append([diag, agg["талонов"], round(float(agg["сумма_талонов"]), 2)])

    ws2 = wb.create_sheet("Талоны 1 услуга")
    ws2.append(
        [
            "Талон",
            "ЕНП",
            "Статус",
            "Подразделение",
            "Врач",
            "Диагноз (DS1)",
            "МКБ основной",
            "Категория (основной)",
            "Сопутствующий (DS2)",
            "МКБ сопутствующие",
            "Категории сопутствующих",
            "Замечания по диагнозам",
            "Услуги (код — сумма)",
            "Сумма талона",
        ]
    )
    for t in sorted(report["single_service"], key=lambda x: x["талон"]):
        ws2.append(
            [
                t["талон"],
                t.get("енп", ""),
                t.get("статус", ""),
                t.get("подразделение", ""),
                t.get("врач", ""),
                t["диагноз"],
                t.get("ds1_code", ""),
                t.get("ds1_category", ""),
                t.get("ds2_raw", ""),
                t.get("ds2_codes", ""),
                t.get("ds2_categories_line", ""),
                t.get("diagnosis_note", ""),
                t.get("услуги", ""),
                round(float(t["сумма_талона"]), 2),
            ]
        )

    ws3 = wb.create_sheet("Талоны 2+ услуг")
    ws3.append(
        [
            "Талон",
            "ЕНП",
            "Статус",
            "Подразделение",
            "Врач",
            "Диагноз (DS1)",
            "МКБ основной",
            "Категория (основной)",
            "Сопутствующий (DS2)",
            "МКБ сопутствующие",
            "Категории сопутствующих",
            "Замечания по диагнозам",
            "Услуги (код — сумма)",
            "Сумма талона",
            "Число строк услуг",
        ]
    )
    for t in sorted(report["multi_service"], key=lambda x: (-x["число_строк_услуг"], x["талон"])):
        ws3.append(
            [
                t["талон"],
                t.get("енп", ""),
                t.get("статус", ""),
                t.get("подразделение", ""),
                t.get("врач", ""),
                t["диагноз"],
                t.get("ds1_code", ""),
                t.get("ds1_category", ""),
                t.get("ds2_raw", ""),
                t.get("ds2_codes", ""),
                t.get("ds2_categories_line", ""),
                t.get("diagnosis_note", ""),
                t.get("услуги", ""),
                round(float(t["сумма_талона"]), 2),
                t["число_строк_услуг"],
            ]
        )

    ws4 = wb.create_sheet("Коды не из справочника")
    inv_header = [
        "Талон",
        "ЕНП",
        "Статус",
        "Подразделение",
        "Врач",
        "Диагноз (DS1)",
        "Код услуги",
        "Причина",
        "Сумма услуги",
    ]
    if report.get("has_ds2_column"):
        inv_header.insert(6, "Сопутствующий (DS2)")
    ws4.append(inv_header)
    for line in report["invalid_lines"]:
        row4 = [
            line["талон"],
            line.get("енп", ""),
            line.get("статус", ""),
            line.get("подразделение", ""),
            line.get("врач", ""),
            line["диагноз"],
            line["код_услуги"],
            line["причина"],
            round(float(line["сумма_услуги"]), 2),
        ]
        if report.get("has_ds2_column"):
            row4.insert(6, line.get("диагноз_ds2", ""))
        ws4.append(row4)

    ws5 = wb.create_sheet("Свод по «левым» кодам")
    ws5.append(["Код услуги", "Строк в выгрузке"])
    for code, n in report["invalid_by_code"].items():
        ws5.append([code, n])

    ws6 = wb.create_sheet("Строки без кода услуги")
    nc_header = [
        "Талон",
        "ЕНП",
        "Статус",
        "Подразделение",
        "Врач",
        "Диагноз (DS1)",
        "Поле «Код услуги»",
        "Сумма услуги",
    ]
    if report.get("has_ds2_column"):
        nc_header.insert(6, "Сопутствующий (DS2)")
    ws6.append(nc_header)
    for line in report["no_code_lines"]:
        row6 = [
            line["талон"],
            line.get("енп", ""),
            line.get("статус", ""),
            line.get("подразделение", ""),
            line.get("врач", ""),
            line["диагноз"],
            line["код_услуги"],
            round(float(line["сумма_услуги"]), 2),
        ]
        if report.get("has_ds2_column"):
            row6.insert(6, line.get("диагноз_ds2", ""))
        ws6.append(row6)

    wsvc = wb.create_sheet("Талоны — проблемы по услугам")
    wsvc.append(
        [
            "Талон",
            "ЕНП",
            "Статус",
            "Подразделение",
            "Врач",
            "МКБ (DS1)",
            "МКБ сопутствующие (DS2)",
            "Услуги по диагнозам (справочник ДН)",
            "Строк услуг",
            "Сумма талона",
            "Причины",
        ]
    )
    for t in sorted(
        report.get("tickets_service_mismatch") or [],
        key=lambda x: x["талон"],
    ):
        wsvc.append(
            [
                t["талон"],
                t.get("енп", ""),
                t.get("статус", ""),
                t.get("подразделение", ""),
                t.get("врач", ""),
                t.get("диагноз", ""),
                t.get("сопутствующие_мкб", ""),
                t.get("услуги_по_диагнозам", ""),
                t["число_строк_услуг"],
                round(float(t["сумма_талона"]), 2),
                t["причины"],
            ]
        )

    if report.get("diagnosis_enriched") and report.get("by_category_ds1"):
        wsc1 = wb.create_sheet("Свод по категориям (DS1)")
        wsc1.append(
            [
                "Категория / статус основного диагноза",
                "Талонов",
                "Сумма талонов",
                "Талонов (услуги по справочнику)",
                "Сумма (услуги по справочнику)",
                "Талонов (услуги не по справочнику)",
                "Сумма (услуги не по справочнику)",
            ]
        )
        for cat, agg in sorted(
            report["by_category_ds1"].items(),
            key=lambda x: (-x[1]["талонов"], x[0]),
        ):
            n = int(agg["талонов"])
            n_ok = int(agg.get("талонов_услуги_ок", 0))
            s = round(float(agg["сумма_талонов"]), 2)
            s_ok = round(float(agg.get("сумма_услуги_ок", 0.0)), 2)
            wsc1.append(
                [
                    cat,
                    n,
                    s,
                    n_ok,
                    s_ok,
                    n - n_ok,
                    round(s - s_ok, 2),
                ]
            )

    if report.get("diagnosis_enriched") and report.get("by_category_ds2"):
        wsc2 = wb.create_sheet("Сопутствующие по категориям")
        wsc2.append(["Категория / статус", "Упоминаний кодов в DS2"])
        for cat, n in report["by_category_ds2"].items():
            wsc2.append([cat, n])

    if report.get("diagnosis_enriched") and report.get("diagnosis_mismatches"):
        wsdx = wb.create_sheet("Диагнозы — несоответствия")
        wsdx.append(
            [
                "Талон",
                "ЕНП",
                "Статус",
                "Подразделение",
                "Врач",
                "Тип",
                "Код МКБ",
                "Комментарий",
            ]
        )
        for m in report["diagnosis_mismatches"]:
            wsdx.append(
                [
                    m["талон"],
                    m.get("енп", ""),
                    m.get("статус", ""),
                    m.get("подразделение", ""),
                    m.get("врач", ""),
                    m["тип"],
                    m["код_мкб"],
                    m["комментарий"],
                ]
            )

    if report.get("diagnosis_enriched") and report.get("by_category_ds1_by_unit"):
        wsc1u = wb.create_sheet("Категории DS1 по подразделениям")
        wsc1u.append(
            [
                "Категория / статус (DS1)",
                "Подразделение",
                "Талонов",
                "Сумма талонов",
                "Талонов (услуги по справочнику)",
                "Сумма (услуги по справочнику)",
                "Талонов (услуги не по справочнику)",
                "Сумма (услуги не по справочнику)",
            ]
        )
        for row in report["by_category_ds1_by_unit"]:
            n = int(row["талонов"])
            n_ok = int(row["талонов_услуги_ок"])
            s = float(row["сумма_талонов"])
            s_ok = float(row["сумма_услуги_ок"])
            wsc1u.append(
                [
                    row["категория"],
                    row["подразделение"],
                    n,
                    round(s, 2),
                    n_ok,
                    round(s_ok, 2),
                    n - n_ok,
                    round(s - s_ok, 2),
                ]
            )

    wb.save(out_path)
